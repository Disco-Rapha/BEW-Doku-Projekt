"""Generiert die synthetischen Test-Fixtures unter ~/Disco-dev/.test-fixtures/.

Idempotent — laeuft beliebig oft, ueberschreibt bestehende synthetische
Files. Files, die manuell aus Prod kommen muessen (Slot 02, 07, 09, 10,
12, 13), werden NICHT angefasst.

Aufruf:
    uv run python scripts/generate_test_fixtures.py
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib import colors

POOL = Path.home() / "Disco-dev" / ".test-fixtures"
SOURCES = POOL / "sources-pool"
CONTEXT = POOL / "context-pool"


def make_01_datenblatt() -> Path:
    """01 Datenblatt: 2-Seiten-A4-PDF mit Header + Tabelle, text-dominant."""
    out = SOURCES / "01_datenblatt.pdf"
    doc = SimpleDocTemplate(str(out), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Datenblatt — Armatur Y0SBD32 AA501", styles["Title"]))
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        "Dieses Datenblatt beschreibt die Armatur Y0SBD32 AA501 nach VGB S 831. "
        "Hersteller: BEW Demonstration GmbH. Druckstufe PN16, Nennweite DN50.",
        styles["Normal"],
    ))
    story.append(Spacer(1, 0.3 * cm))

    # Datentabelle
    data = [
        ["Parameter", "Wert", "Einheit"],
        ["Druckstufe", "PN16", "bar"],
        ["Nennweite", "DN50", "mm"],
        ["Werkstoff Gehäuse", "EN-JL1040", "—"],
        ["Werkstoff Sitz", "X20Cr13", "—"],
        ["Maximale Temperatur", "180", "°C"],
        ["Antriebsart", "elektrisch", "—"],
        ["Schutzart", "IP65", "—"],
    ]
    t = Table(data, colWidths=[6 * cm, 5 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5 * cm))

    # Seite 2
    story.append(Paragraph("Wartungsanforderungen", styles["Heading2"]))
    story.append(Paragraph(
        "Sichtprüfung jährlich. Funktionstest alle 6 Monate gemäß Wartungsplan. "
        "Bei Auffälligkeiten Hersteller-Spezifikation Abschnitt 4.2 konsultieren.",
        styles["Normal"],
    ))
    story.append(Paragraph(
        "Druckprüfung nach VGB-S-831 Anhang B mindestens alle 4 Jahre, sofern "
        "im Anlagenbetriebshandbuch nicht anders festgelegt.",
        styles["Normal"],
    ))
    doc.build(story)
    return out


def make_03_scan_protokoll() -> Path:
    """03 Scan-Protokoll: PIL erzeugt Schreibmaschinen-Look JPG, in PDF eingebettet."""
    # Schreibmaschinen-Bild
    img = Image.new("RGB", (1240, 1754), color="white")  # A4 @ 150 DPI
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Courier.dfont", 28)
        font_small = ImageFont.truetype("/System/Library/Fonts/Courier.dfont", 22)
    except Exception:
        font = ImageFont.load_default()
        font_small = font
    draw.text((100, 100), "PROTOKOLL — INBETRIEBNAHME-PRÜFUNG", fill="black", font=font)
    draw.text((100, 160), "=" * 50, fill="black", font=font_small)
    draw.text((100, 220), "Anlage:    Lagerhalle BEW Reuter", fill="black", font=font_small)
    draw.text((100, 260), "Gewerk:    Mechanik / Armaturen", fill="black", font=font_small)
    draw.text((100, 300), "Datum:     2026-04-15", fill="black", font=font_small)
    draw.text((100, 340), "Prüfer:    M. Schmidt (BEW QM)", fill="black", font=font_small)
    draw.text((100, 420), "Befund:", fill="black", font=font)
    draw.text((100, 480), "1. Sichtprüfung Y0SBD32 AA501 — i.O.", fill="black", font=font_small)
    draw.text((100, 520), "2. Funktionstest Stellantrieb — i.O.", fill="black", font=font_small)
    draw.text((100, 560), "3. Druckprüfung 1.5 × PN — bestanden", fill="black", font=font_small)
    draw.text((100, 600), "4. Dokumentation laut VGB-S-831 — vollständig", fill="black", font=font_small)
    img_temp = SOURCES / "_03_scan_temp.png"
    img.save(img_temp)

    out = SOURCES / "03_scan_protokoll.pdf"
    c = canvas.Canvas(str(out), pagesize=A4)
    c.drawImage(str(img_temp), 0, 0, width=A4[0], height=A4[1])
    c.save()
    img_temp.unlink()
    return out


def make_04_kks_schild() -> Path:
    """04 KKS-Schild als JPG."""
    out = SOURCES / "04_kks_schild.jpg"
    img = Image.new("RGB", (800, 600), color=(220, 220, 220))
    draw = ImageDraw.Draw(img)
    try:
        font_big = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 90)
        font_small = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 30)
    except Exception:
        font_big = ImageFont.load_default()
        font_small = font_big
    # Rahmen
    draw.rectangle([(20, 20), (780, 580)], outline="black", width=4)
    draw.text((100, 80), "Y0SBD32", fill="black", font=font_big)
    draw.text((100, 200), "AA501", fill="black", font=font_big)
    draw.text((100, 360), "Armatur", fill="black", font=font_small)
    draw.text((100, 410), "PN16 / DN50", fill="black", font=font_small)
    draw.text((100, 460), "BEW Lagerhalle", fill="black", font=font_small)
    img.save(out, "JPEG", quality=85)
    return out


def make_05_lieferindex() -> Path:
    """05 Lieferindex: 2-Sheet-Excel."""
    out = SOURCES / "05_lieferindex.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Geräteliste"
    headers = ["KKS-Code", "Bezeichnung", "Hersteller", "Typ", "DN", "PN"]
    ws1.append(headers)
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    rows = [
        ["Y0SBD32 AA501", "Schiebearmatur Vorlauf", "BEW Demo GmbH", "DN50-PN16-EL", 50, 16],
        ["Y0SBD32 AA502", "Schiebearmatur Rücklauf", "BEW Demo GmbH", "DN50-PN16-EL", 50, 16],
        ["Y0SBD31 AA501", "Absperrventil Heizkreis 1", "ARI Armaturen", "ZESA-DN40", 40, 16],
        ["Y0SBD31 AA502", "Absperrventil Heizkreis 2", "ARI Armaturen", "ZESA-DN40", 40, 16],
        ["Y0SBD31 CP501", "Druckmessumformer", "WIKA", "S-20", 25, 25],
    ]
    for row in rows:
        ws1.append(row)

    ws2 = wb.create_sheet("Status")
    ws2.append(["KKS-Code", "Lieferstatus", "Datum", "Bemerkung"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    statuses = [
        ["Y0SBD32 AA501", "geliefert", "2026-03-15", ""],
        ["Y0SBD32 AA502", "geliefert", "2026-03-15", ""],
        ["Y0SBD31 AA501", "ausstehend", "—", "verzögert wegen Werkstoff-Engpass"],
        ["Y0SBD31 AA502", "ausstehend", "—", "—"],
        ["Y0SBD31 CP501", "geliefert", "2026-04-02", "—"],
    ]
    for row in statuses:
        ws2.append(row)

    wb.save(out)
    return out


def make_06_dcc_katalog() -> Path:
    """06 DCC-Katalog: 1-Sheet Excel mit Lookup-Struktur (für context/)."""
    out = CONTEXT / "06_dcc_katalog.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DCC"
    ws.append(["DCC-Code", "Bezeichnung", "Beschreibung", "Pflicht-Doku"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    rows = [
        ["A.1.1", "Anlagenübersicht", "Übersichts-Schema mit Hauptkomponenten", "ja"],
        ["A.2.1", "Systemzuordnung", "Zuordnung KKS-Code zu Anlagensystem", "ja"],
        ["A.2.2", "RDS-PP-Zuordnung", "Bezeichnung nach RDS-PP für jeden Komponenten-Typ", "ja"],
        ["A.3.1", "Datenblätter", "Hersteller-Datenblätter pro Komponente", "ja"],
        ["A.3.2", "Materialzeugnisse", "Werkstoff-Zertifikate nach EN 10204", "ja"],
        ["A.4.1", "Schaltpläne", "Elektrische Schalt- und Anschlusspläne", "ja"],
        ["A.5.1", "Inbetriebnahme-Protokolle", "Endkontrollen + Funktionsprüfungen", "ja"],
        ["A.5.2", "Druckprüfungen", "Druck-/Dichtheitsprüfungs-Protokolle", "ja"],
        ["B.1.1", "Wartungsanleitungen", "Hersteller-Wartungspläne", "ja"],
        ["B.2.1", "Ersatzteilliste", "Empfohlene Ersatzteile pro Anlagenteil", "nein"],
    ]
    for row in rows:
        ws.append(row)
    wb.save(out)
    return out


def make_08_leeres_pdf() -> Path:
    """08 Leeres PDF: 1 Seite, kein Text."""
    out = SOURCES / "08_leeres_dokument.pdf"
    c = canvas.Canvas(str(out), pagesize=A4)
    # Eine leere Seite — keine Content-Calls
    c.showPage()
    c.save()
    return out


def make_11_duplikat() -> Path:
    """11 Duplikat: byte-identische Kopie von 01."""
    src = SOURCES / "01_datenblatt.pdf"
    out = SOURCES / "11_duplikat_von_01.pdf"
    if src.exists():
        shutil.copy2(src, out)
    return out


def make_14_bericht_versionen() -> tuple[Path, Path]:
    """14a/14b: zwei Bericht-Versionen mit gleichem Stamm.

    14a ist 1 Page (kurz), 14b ist 3 Pages (lang, ergänzte Inhalte).
    Beide haben Inhalte die der gleichen Sache thematisch zugeordnet sind,
    sodass die replaces-Heuristik (wenn implementiert) sie als
    Versionskette erkennen kann.
    """
    a = SOURCES / "14a_bericht_kurzversion.pdf"
    b = SOURCES / "14b_bericht_langversion.pdf"
    styles = getSampleStyleSheet()

    # 14a — kurz
    doc_a = SimpleDocTemplate(str(a), pagesize=A4)
    story_a = [
        Paragraph("Bericht — Inbetriebnahme Lagerhalle (Stand 2026-03)", styles["Title"]),
        Spacer(1, 0.5 * cm),
        Paragraph(
            "Erste Inbetriebnahme abgeschlossen. Detail-Bericht folgt mit "
            "Status der ausstehenden Komponenten.",
            styles["Normal"],
        ),
    ]
    doc_a.build(story_a)

    # 14b — lang (gleicher Stamm, ergänzt)
    doc_b = SimpleDocTemplate(str(b), pagesize=A4)
    story_b = [
        Paragraph("Bericht — Inbetriebnahme Lagerhalle (Stand 2026-04, Endversion)", styles["Title"]),
        Spacer(1, 0.5 * cm),
        Paragraph(
            "Inbetriebnahme vollständig abgeschlossen einschließlich der zuvor "
            "ausstehenden Heizkreis-Absperrventile.",
            styles["Normal"],
        ),
        Spacer(1, 0.3 * cm),
        Paragraph("Status pro Komponente", styles["Heading2"]),
        Paragraph(
            "Y0SBD32 AA501/502: geliefert + funktionsgeprüft. "
            "Y0SBD31 AA501/502: nachgeliefert 2026-04-08, integriert + geprüft. "
            "Y0SBD31 CP501: geliefert + kalibriert.",
            styles["Normal"],
        ),
        Spacer(1, 0.3 * cm),
        Paragraph("Druckprüfung", styles["Heading2"]),
        Paragraph(
            "Hauptkreis 1.5 × PN bestanden. Heizkreis-Schleife getrennt geprüft, "
            "ohne Beanstandungen. Protokolle in Anhang.",
            styles["Normal"],
        ),
        Spacer(1, 0.3 * cm),
        Paragraph("Anhänge", styles["Heading2"]),
        Paragraph("A1 Druckprüfprotokoll (separates Dokument)", styles["Normal"]),
        Paragraph("A2 Funktionsprüfung Antriebe (separates Dokument)", styles["Normal"]),
    ]
    doc_b.build(story_b)
    return a, b


def main() -> None:
    print(f"Pool-Verzeichnis: {POOL}")
    if not POOL.is_dir():
        raise SystemExit(f"Pool-Verzeichnis fehlt: {POOL}. Erst `mkdir -p` aufrufen.")

    SOURCES.mkdir(exist_ok=True)
    CONTEXT.mkdir(exist_ok=True)

    generators = [
        ("01_datenblatt.pdf", make_01_datenblatt),
        ("03_scan_protokoll.pdf", make_03_scan_protokoll),
        ("04_kks_schild.jpg", make_04_kks_schild),
        ("05_lieferindex.xlsx", make_05_lieferindex),
        ("06_dcc_katalog.xlsx", make_06_dcc_katalog),
        ("08_leeres_dokument.pdf", make_08_leeres_pdf),
        ("11_duplikat_von_01.pdf", make_11_duplikat),
        ("14a/14b bericht_versionen", make_14_bericht_versionen),
    ]

    for label, fn in generators:
        try:
            result = fn()
            if isinstance(result, tuple):
                for p in result:
                    print(f"  ✓ {p.name} ({p.stat().st_size} bytes)")
            else:
                print(f"  ✓ {result.name} ({result.stat().st_size} bytes)")
        except Exception as exc:
            print(f"  ✗ {label} — FEHLER: {exc}")

    print()
    print("Manuell zu beschaffende Files (siehe MANIFEST.md):")
    for slot in ["02_schaltplan_a3.pdf", "07_grundriss.dwg",
                 "09_korruptes_dokument.pdf", "10_korruptes_zeichnung.dwg",
                 "12_bericht.docx", "13_praesentation.pptx"]:
        path = SOURCES / slot
        present = "✓ vorhanden" if path.exists() else "❌ fehlt — bitte rein in den Pool"
        print(f"  {slot:35s}  {present}")


if __name__ == "__main__":
    main()
