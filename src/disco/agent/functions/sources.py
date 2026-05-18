"""Sources-Registry: rekursiver Scan von sources/ und/oder context/ + Begleit-Metadaten + Duplikat-Erkennung.

Pattern: "File Registry with Change Detection"
  - sha256-basiert fuer sichere Aenderungserkennung
  - Idempotent: wiederholter Lauf auf unveraendertem Filesystem ergibt 0 Delta
  - Status-Feld 'active'|'deleted' fuer Abgang-Erkennung
  - `kind`-Tag ('source' vs. 'context') trennt Arbeitsdokumente von
    Arbeitsgrundlagen in derselben Pipeline

Performance: auf SSD ca. 2-5s fuer 1600 Dateien bei gemischten Groessen.
Hash wird gecacht per (rel_path, size, mtime) — wenn nichts davon sich
geaendert hat, ueberspringen wir das erneute Hashen (optimierung).

Pipeline-Bridge (Stufe 1 Architektur):
  Nach jedem Scan werden alle aktiven PDFs automatisch in die PDF-Pipeline-
  Einstiegstabelle `agent_pdf_inventory` (Ebene 2, datastore.db) gespiegelt.
  Damit ist der Standard-Flow `Registrieren → Routing → Extraktion` ohne
  manuellen Zwischenschritt lauffaehig — `pdf_routing_decision` findet
  sofort gefuelltes Inventar vor.

Scope-Konvention:
  agent_sources.rel_path ist **relativ zum jeweiligen Scope-Root**
  (zu sources/ bzw. zu context/). Der Projekt-Root-relative Pfad wird beim
  PDF-Inventory-Sync zusammengebaut (z.B. "sources/foo.pdf" oder
  "context/norm.pdf"), damit die PDF-Pipeline-Flows einheitliche Pfade
  sehen.
"""

from __future__ import annotations

import hashlib
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from . import register
from ..context import connect_datastore_rw as _connect_datastore_rw
from .fs import _data_root


# Ordner/Namen die niemals gescannt werden (Metadaten-Container)
SCAN_IGNORE_TOP: frozenset[str] = frozenset({"_meta", ".DS_Store"})
SCAN_IGNORE_SUFFIX: tuple[str, ...] = (".tmp", ".crdownload", ".part")
SCAN_IGNORE_PATTERNS: tuple[str, ...] = (".git", "__pycache__", "node_modules")


def _sha256_file(path: Path, chunk_size: int = 1 << 20) -> str:
    """SHA-256 ueber den Datei-Inhalt, chunk-basiert (Memory-safe bei grossen PDFs)."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _iso_mtime(epoch: float) -> str:
    return datetime.fromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S")


def _is_ignored(rel_path: Path) -> bool:
    """True wenn dieser Pfad beim Scan uebersprungen wird.

    Konvention (siehe CLAUDE.md): Pfad-Parts mit '_'- oder '.'-Prefix sind
    INTERN und werden generell ignoriert — egal ob Ordner (_meta/, .git/)
    oder Datei (_manifest.md, .DS_Store, .gitignore). Das haelt die
    Registry konsistent mit dem Pipeline-Status-Filter, der dieselbe
    Konvention anwendet.
    """
    parts = rel_path.parts
    # '_'- oder '.'-Prefix in irgendeinem Pfad-Part: intern, ignorieren
    for p in parts:
        if p.startswith("_") or p.startswith("."):
            return True
    # Sonder-Pattern irgendwo im Pfad (zusaetzlich zu Prefix-Filter)
    for p in parts:
        if p in SCAN_IGNORE_PATTERNS:
            return True
    # Suffix-Pattern
    for suf in SCAN_IGNORE_SUFFIX:
        if rel_path.name.endswith(suf):
            return True
    return False


# Mapping scope → (kind, folder_name_under_project_root)
_SCOPE_TO_KIND: dict[str, tuple[str, str]] = {
    "sources": ("source", "sources"),
    "context": ("context", "context"),
}


def _scan_one_scope(
    *,
    conn,
    scan_id: int,
    kind: str,
    scope_root: Path,
    subpath: str,
    skip_hash_if_unchanged: bool,
) -> dict[str, Any]:
    """Scannt einen Unterbaum (sources/ ODER context/), aktualisiert agent_sources.

    rel_path wird relativ zum `scope_root` gespeichert (wie bisher fuer sources/).
    Der `kind`-Tag unterscheidet die beiden Welten.

    Returns: {"new": [...], "changed": [...], "deleted": [...], "unchanged": int}
    """
    scan_subdir = (scope_root / subpath).resolve() if subpath else scope_root
    try:
        scan_subdir.relative_to(scope_root)
    except ValueError:
        raise ValueError(
            f"subpath ausserhalb von {scope_root.name}/: {subpath!r}"
        )

    # Bestehende aktive/deleted-Eintraege dieses kind-Scopes laden
    if subpath:
        like = f"{subpath.rstrip('/')}/%"
        rows = conn.execute(
            "SELECT id, rel_path, size_bytes, sha256, mtime, status "
            "FROM agent_sources "
            "WHERE kind = ? AND (rel_path LIKE ? OR rel_path = ?)",
            (kind, like, subpath.rstrip("/")),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, rel_path, size_bytes, sha256, mtime, status "
            "FROM agent_sources WHERE kind = ?",
            (kind,),
        ).fetchall()
    existing: dict[str, dict] = {r["rel_path"]: dict(r) for r in rows}

    seen_rel_paths: set[str] = set()
    new_list: list[str] = []
    changed_list: list[str] = []
    unchanged_count = 0

    for fs_path in scan_subdir.rglob("*"):
        if not fs_path.is_file():
            continue
        rel = fs_path.relative_to(scope_root)
        if _is_ignored(rel):
            continue
        rel_str = str(rel)
        seen_rel_paths.add(rel_str)

        try:
            st = fs_path.stat()
        except OSError:
            continue
        size = st.st_size
        mtime_iso = _iso_mtime(st.st_mtime)
        folder = str(rel.parent) if rel.parent != Path(".") else ""
        ext = rel.suffix.lstrip(".").lower() or None

        prev = existing.get(rel_str)

        need_hash = True
        if prev and skip_hash_if_unchanged:
            if prev["size_bytes"] == size and prev["mtime"] == mtime_iso:
                need_hash = False

        digest = None
        if need_hash:
            try:
                digest = _sha256_file(fs_path)
            except OSError as exc:
                conn.execute(
                    "UPDATE agent_source_scans SET notes = "
                    "COALESCE(notes, '') || ? WHERE id = ?",
                    (f"Lesefehler [{kind}] {rel_str}: {exc}\n", scan_id),
                )
                continue
        else:
            digest = prev["sha256"]

        if not prev:
            # canonical_path direkt beim Insert berechnen (Migration 011).
            # PathResolver kanonisiert NFD→NFC + ' : '→'/' auf macOS.
            from disco.fs.path_resolver import get_resolver
            canonical = get_resolver().to_canonical(rel_str)
            conn.execute(
                "INSERT INTO agent_sources "
                "(rel_path, canonical_path, filename, folder, extension, size_bytes, sha256, "
                " mtime, kind, first_seen_at, last_seen_at, last_changed_at, status) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'), "
                "        datetime('now'), 'active')",
                (rel_str, canonical, fs_path.name, folder, ext, size, digest, mtime_iso, kind),
            )
            new_list.append(rel_str)
        else:
            hash_changed = (prev["sha256"] or "") != (digest or "")
            was_deleted = prev["status"] == "deleted"
            if hash_changed:
                conn.execute(
                    "UPDATE agent_sources SET "
                    "  size_bytes = ?, sha256 = ?, mtime = ?, "
                    "  last_seen_at = datetime('now'), "
                    "  last_changed_at = datetime('now'), "
                    "  status = 'active', "
                    "  updated_at = datetime('now') "
                    "WHERE id = ?",
                    (size, digest, mtime_iso, prev["id"]),
                )
                changed_list.append(rel_str)
            else:
                if was_deleted:
                    conn.execute(
                        "UPDATE agent_sources SET "
                        "  status = 'active', last_seen_at = datetime('now'), "
                        "  last_changed_at = datetime('now'), "
                        "  updated_at = datetime('now') "
                        "WHERE id = ?",
                        (prev["id"],),
                    )
                    changed_list.append(rel_str)
                else:
                    conn.execute(
                        "UPDATE agent_sources SET "
                        "  last_seen_at = datetime('now') "
                        "WHERE id = ?",
                        (prev["id"],),
                    )
                    unchanged_count += 1

    # Abgang-Erkennung: Eintraege die in DB fuer diesen Scope existieren,
    # aber nicht mehr im FS — als 'deleted' markieren.
    # ACHTUNG: Bei subpath nur innerhalb des Subpath, sonst Gefahr dass
    # wir Dateien als 'deleted' flaggen die im nicht-gescannten Zweig liegen.
    deleted_list: list[str] = []
    for rel_str, prev in existing.items():
        if rel_str in seen_rel_paths:
            continue
        if prev["status"] != "active":
            continue
        # Wenn subpath gegeben: nur Kandidaten unter diesem subpath
        if subpath:
            prefix = subpath.rstrip("/") + "/"
            if rel_str != subpath.rstrip("/") and not rel_str.startswith(prefix):
                continue
        conn.execute(
            "UPDATE agent_sources SET "
            "  status = 'deleted', "
            "  last_changed_at = datetime('now'), "
            "  updated_at = datetime('now') "
            "WHERE id = ?",
            (prev["id"],),
        )
        deleted_list.append(rel_str)

    return {
        "kind": kind,
        "new": new_list,
        "changed": changed_list,
        "deleted": deleted_list,
        "unchanged": unchanged_count,
    }


def _scan_one_scope_v2(
    *,
    conn,
    scan_id: int,
    kind: str,
    scope_root: Path,
    subpath: str,
    skip_hash_if_unchanged: bool,
) -> dict[str, Any]:
    """Hash-zentrierter Scan (Pipeline-Reform v2).

    Schreibt agent_sources (pro Hash) + agent_source_locations (pro Pfad).
    Cases die behandelt werden:
      - C1  Neue Datei (neuer Hash)           → INSERT source + location
      - C1' Neue Location, Hash schon da      → INSERT nur location
      - C2  Move/Rename (gleicher Hash)       → wird hier als delete-alt + new-pfad
              behandelt; rel_path-UPDATE-Pfad braucht eine separate Heuristik
      - C4  Neue Version (gleicher Pfad, neuer Hash) → alte location 'deleted',
              neue location mit neuer source_id
      - C5  Duplikat                          → C1' analog
      - C6  Datei gelöscht (verschwunden)     → location 'deleted', source ggf. 'deleted'
      - CR  Wiederauferstehung                → status 'deleted' → 'active'
    """
    scan_subdir = (scope_root / subpath).resolve() if subpath else scope_root
    try:
        scan_subdir.relative_to(scope_root)
    except ValueError:
        raise ValueError(
            f"subpath ausserhalb von {scope_root.name}/: {subpath!r}"
        )

    # 1. Bestehende Locations dieses kinds laden (mit Source-Info)
    if subpath:
        like = f"{subpath.rstrip('/')}/%"
        rows = conn.execute(
            "SELECT l.id AS loc_id, l.rel_path, l.status AS loc_status, "
            "       l.mtime, l.source_id, s.sha256, s.size_bytes, "
            "       s.status AS src_status "
            "FROM agent_source_locations l "
            "JOIN agent_sources s ON s.id = l.source_id "
            "WHERE s.kind = ? AND (l.rel_path LIKE ? OR l.rel_path = ?)",
            (kind, like, subpath.rstrip("/")),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT l.id AS loc_id, l.rel_path, l.status AS loc_status, "
            "       l.mtime, l.source_id, s.sha256, s.size_bytes, "
            "       s.status AS src_status "
            "FROM agent_source_locations l "
            "JOIN agent_sources s ON s.id = l.source_id "
            "WHERE s.kind = ?",
            (kind,),
        ).fetchall()
    existing_by_path: dict[str, dict] = {r["rel_path"]: dict(r) for r in rows}

    # 2. Hash-Index aller Sources dieses kinds (auch deleted, weil
    #    Wiederauferstehung möglich)
    src_by_hash: dict[str, int] = {}
    for r in conn.execute(
        "SELECT id, sha256 FROM agent_sources WHERE kind = ? AND sha256 IS NOT NULL",
        (kind,),
    ).fetchall():
        src_by_hash[r["sha256"]] = int(r["id"])

    seen_loc_ids: set[int] = set()
    new_list: list[str] = []
    changed_list: list[str] = []
    unchanged_count = 0

    from disco.fs.path_resolver import get_resolver
    resolver = get_resolver()

    for fs_path in scan_subdir.rglob("*"):
        if not fs_path.is_file():
            continue
        rel = fs_path.relative_to(scope_root)
        if _is_ignored(rel):
            continue
        rel_str = str(rel)

        try:
            st = fs_path.stat()
        except OSError:
            continue
        size = st.st_size
        mtime_iso = _iso_mtime(st.st_mtime)
        folder = str(rel.parent) if rel.parent != Path(".") else ""
        ext = rel.suffix.lstrip(".").lower() or None
        canonical = resolver.to_canonical(rel_str)

        prev_loc = existing_by_path.get(rel_str)

        # Hash berechnen (oder cachen über mtime+size)
        need_hash = True
        if prev_loc and skip_hash_if_unchanged:
            if prev_loc["size_bytes"] == size and prev_loc["mtime"] == mtime_iso:
                need_hash = False
        if need_hash:
            try:
                digest = _sha256_file(fs_path)
            except OSError as exc:
                conn.execute(
                    "UPDATE agent_source_scans SET notes = "
                    "COALESCE(notes, '') || ? WHERE id = ?",
                    (f"Lesefehler [{kind}] {rel_str}: {exc}\n", scan_id),
                )
                continue
        else:
            digest = prev_loc["sha256"]

        # 3. Source finden oder anlegen
        source_id = src_by_hash.get(digest)
        if source_id is None:
            cur = conn.execute(
                "INSERT INTO agent_sources "
                "(sha256, size_bytes, kind, status, first_seen_at, last_seen_at, "
                " last_changed_at, created_at, updated_at) "
                "VALUES (?, ?, ?, 'active', datetime('now'), datetime('now'), "
                "        datetime('now'), datetime('now'), datetime('now'))",
                (digest, size, kind),
            )
            source_id = int(cur.lastrowid)
            src_by_hash[digest] = source_id
        else:
            # Wiederauferstehung der Source, falls vorher 'deleted'
            conn.execute(
                "UPDATE agent_sources SET status='active', "
                "  last_seen_at=datetime('now'), updated_at=datetime('now') "
                "WHERE id = ? AND status != 'active'",
                (source_id,),
            )

        # 4. Location handhaben
        if prev_loc is None:
            # C1 oder C1' — neue Location
            cur = conn.execute(
                "INSERT INTO agent_source_locations "
                "(source_id, rel_path, origin, status, first_seen_at, last_seen_at, "
                " mtime, filename, folder, extension, canonical_path, "
                " created_at, updated_at) "
                "VALUES (?, ?, 'local-folder', 'active', datetime('now'), "
                "        datetime('now'), ?, ?, ?, ?, ?, "
                "        datetime('now'), datetime('now'))",
                (source_id, rel_str, mtime_iso, fs_path.name, folder, ext, canonical),
            )
            seen_loc_ids.add(int(cur.lastrowid))
            new_list.append(rel_str)
        elif prev_loc["source_id"] == source_id:
            # Bekannte Location, gleicher Hash
            if prev_loc["loc_status"] == "deleted":
                # CR Wiederauferstehung
                conn.execute(
                    "UPDATE agent_source_locations SET "
                    "  status='active', last_seen_at=datetime('now'), "
                    "  mtime=?, updated_at=datetime('now') WHERE id = ?",
                    (mtime_iso, prev_loc["loc_id"]),
                )
                seen_loc_ids.add(int(prev_loc["loc_id"]))
                changed_list.append(rel_str)
            else:
                # Aktiv, unverändert — nur last_seen aktualisieren
                conn.execute(
                    "UPDATE agent_source_locations SET "
                    "  last_seen_at=datetime('now') WHERE id = ?",
                    (prev_loc["loc_id"],),
                )
                seen_loc_ids.add(int(prev_loc["loc_id"]))
                unchanged_count += 1
        else:
            # C4 — gleicher Pfad, anderer Hash → alte Location 'deleted',
            # neue Location mit neuer source_id (delete + new)
            conn.execute(
                "UPDATE agent_source_locations SET "
                "  status='deleted', last_seen_at=datetime('now'), "
                "  updated_at=datetime('now') WHERE id = ?",
                (prev_loc["loc_id"],),
            )
            cur = conn.execute(
                "INSERT INTO agent_source_locations "
                "(source_id, rel_path, origin, status, first_seen_at, last_seen_at, "
                " mtime, filename, folder, extension, canonical_path, "
                " created_at, updated_at) "
                "VALUES (?, ?, 'local-folder', 'active', datetime('now'), "
                "        datetime('now'), ?, ?, ?, ?, ?, "
                "        datetime('now'), datetime('now'))",
                (source_id, rel_str, mtime_iso, fs_path.name, folder, ext, canonical),
            )
            seen_loc_ids.add(int(cur.lastrowid))
            changed_list.append(rel_str)

    # 5. Vermisste Locations: aktive Einträge, die nicht im FS-Scan auftauchten
    deleted_list: list[str] = []
    for rel_str, prev in existing_by_path.items():
        if int(prev["loc_id"]) in seen_loc_ids:
            continue
        if prev["loc_status"] != "active":
            continue
        if subpath:
            prefix = subpath.rstrip("/") + "/"
            if rel_str != subpath.rstrip("/") and not rel_str.startswith(prefix):
                continue
        conn.execute(
            "UPDATE agent_source_locations SET "
            "  status='deleted', last_seen_at=datetime('now'), "
            "  updated_at=datetime('now') WHERE id = ?",
            (prev["loc_id"],),
        )
        deleted_list.append(rel_str)

    # 6. Sources ohne aktive Locations → status='deleted' (Soft-Delete)
    conn.execute(
        "UPDATE agent_sources SET status='deleted', updated_at=datetime('now') "
        "WHERE kind = ? AND status='active' "
        "AND NOT EXISTS ("
        "  SELECT 1 FROM agent_source_locations l "
        "  WHERE l.source_id = agent_sources.id AND l.status='active'"
        ")",
        (kind,),
    )

    return {
        "kind": kind,
        "new": new_list,
        "changed": changed_list,
        "deleted": deleted_list,
        "unchanged": unchanged_count,
    }


def _sync_pdf_inventory_v2(conn, kinds: list[str]) -> dict[str, int]:
    """Synct agent_source_locations (PDFs) → agent_pdf_inventory (Hash-Modell).

    Strategie: Clean-Rebuild der inventory pro touched kind. agent_pdf_inventory
    ist 1:1 zu Hash-Sources mit aktiver PDF-Location und hat keine eigene State
    (keine externen FK-Referenzen). Daher ist DELETE+INSERT sicher und vermeidet
    Probleme mit stale IDs aus der Pre-Migration-Welt.
    """
    if not kinds:
        total = conn.execute("SELECT COUNT(*) FROM agent_pdf_inventory").fetchone()[0]
        return {"inserted": 0, "updated": 0, "removed": 0,
                "total_in_inventory": int(total)}

    placeholders = ",".join("?" * len(kinds))

    # 1) Alle Inventory-Zeilen für betroffene kinds wegwerfen
    cur = conn.execute(
        f"DELETE FROM agent_pdf_inventory WHERE kind IN ({placeholders})",
        tuple(kinds),
    )
    removed = cur.rowcount or 0

    # 2) PDF-Sources mit mind. einer aktiven Location → Inventory neu befüllen
    rows = conn.execute(
        f"""
        SELECT s.id, s.sha256, s.size_bytes, s.kind,
               (SELECT l.rel_path FROM agent_source_locations l
                WHERE l.source_id = s.id AND l.status='active'
                  AND LOWER(l.extension) = 'pdf'
                ORDER BY l.id LIMIT 1) AS rel_path,
               (SELECT l.filename FROM agent_source_locations l
                WHERE l.source_id = s.id AND l.status='active'
                  AND LOWER(l.extension) = 'pdf'
                ORDER BY l.id LIMIT 1) AS filename,
               (SELECT l.folder FROM agent_source_locations l
                WHERE l.source_id = s.id AND l.status='active'
                  AND LOWER(l.extension) = 'pdf'
                ORDER BY l.id LIMIT 1) AS folder
        FROM agent_sources s
        WHERE s.status = 'active'
          AND s.kind IN ({placeholders})
          AND EXISTS (
            SELECT 1 FROM agent_source_locations l
            WHERE l.source_id = s.id AND l.status='active'
              AND LOWER(l.extension) = 'pdf'
          )
        """,
        tuple(kinds),
    ).fetchall()

    inserted = 0
    for r in rows:
        rel = r["rel_path"] or ""
        # Projekt-relativer Pfad mit sources/- oder context/-Praefix
        role_prefix = "context" if r["kind"] == "context" else "sources"
        full_rel = rel if rel.startswith(f"{role_prefix}/") else f"{role_prefix}/{rel}"
        file_name = r["filename"] or Path(full_rel).name
        gewerk = (r["folder"] or "").split("/")[0] if r["folder"] else None

        try:
            conn.execute(
                "INSERT INTO agent_pdf_inventory "
                "(id, rel_path, file_name, file_name_norm, gewerk, size_bytes, sha256, kind) "
                "VALUES (?, ?, ?, LOWER(TRIM(?)), ?, ?, ?, ?)",
                (r["id"], full_rel, file_name, file_name, gewerk,
                 r["size_bytes"], r["sha256"], r["kind"]),
            )
            inserted += 1
        except Exception:
            # Defensive: bei UNIQUE-Konflikten (z.B. doppelter rel_path durch
            # Edge-Case) skippen — der erste Eintrag pro Pfad gewinnt.
            continue

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM agent_pdf_inventory").fetchone()[0]
    return {"inserted": inserted, "updated": 0, "removed": removed,
            "total_in_inventory": int(total)}


def _sync_pdf_inventory(conn, kinds: list[str]) -> dict[str, int]:
    """Synct agent_sources → agent_pdf_inventory fuer PDFs der angegebenen kinds.

    Baut den Projekt-Root-relativen Pfad ("sources/..." bzw. "context/...") und
    schreibt Upserts nach agent_pdf_inventory. Auch Deletes werden gespiegelt:
    wenn eine Source auf status='deleted' gesetzt ist, wird der Inventory-
    Eintrag entfernt (die Pipeline-Flows sollen sie dann nicht mehr sehen).

    Returns: {"inserted": N, "updated": N, "removed": N}
    """
    if not kinds:
        total_in_inventory = conn.execute(
            "SELECT COUNT(*) FROM agent_pdf_inventory"
        ).fetchone()[0]
        return {
            "inserted": 0,
            "updated": 0,
            "removed": 0,
            "total_in_inventory": int(total_in_inventory),
        }

    placeholders = ",".join("?" * len(kinds))
    # 1) PDFs aus agent_sources holen, die 'active' sind
    src_rows = conn.execute(
        f"""
        SELECT id, rel_path, filename, sha256, size_bytes, folder, kind
        FROM agent_sources
        WHERE status = 'active'
          AND kind IN ({placeholders})
          AND lower(extension) = 'pdf'
        """,
        tuple(kinds),
    ).fetchall()

    inserted = 0
    updated = 0
    for r in src_rows:
        # agent_sources.rel_path ist scope-relativ ("foo.pdf" oder "Elektro/bar.pdf").
        # Fuer die Inventory brauchen wir den Projekt-Root-relativen Pfad, damit
        # Path(rel_path) in den Flow-Runnern direkt funktioniert.
        scope_prefix = "sources" if r["kind"] == "source" else "context"
        inv_rel_path = f"{scope_prefix}/{r['rel_path']}"
        file_name = r["filename"]
        file_name_norm = (file_name or "").strip().lower()

        # Gewerk-Heuristik: erstes Pfad-Segment unter dem scope-Root,
        # wenn es eins gibt (sonst NULL).
        folder = r["folder"] or ""
        gewerk = folder.split("/", 1)[0] if folder else None

        # Upsert via INSERT OR REPLACE auf UNIQUE(rel_path).
        # Auto-Inkrement-ID bleibt stabil solange rel_path unveraendert bleibt,
        # aber wir nutzen REPLACE → deshalb ist file_id NICHT stabil, wenn ein
        # rel_path umbenannt wird. Fuer reine Updates auf gleichem rel_path
        # ist das ok.
        #
        # Wichtig: Wir nutzen ON CONFLICT(rel_path) DO UPDATE, damit wir den
        # bestehenden file_id (agent_pdf_inventory.id) *nicht* verlieren —
        # sonst wuerden work_pdf_routing und agent_pdf_markdown ins Leere
        # referenzieren.
        cur = conn.execute(
            """
            INSERT INTO agent_pdf_inventory
                (rel_path, file_name, file_name_norm, gewerk, size_bytes, sha256, kind)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(rel_path) DO UPDATE SET
                file_name = excluded.file_name,
                file_name_norm = excluded.file_name_norm,
                gewerk = excluded.gewerk,
                size_bytes = excluded.size_bytes,
                sha256 = excluded.sha256,
                kind = excluded.kind
            """,
            (
                inv_rel_path,
                file_name,
                file_name_norm,
                gewerk,
                r["size_bytes"],
                r["sha256"],
                r["kind"],
            ),
        )
        if cur.rowcount > 0:
            # SQLite unterscheidet INSERT vs UPDATE nicht direkt in rowcount
            # beim UPSERT. Heuristik: existierte vorher schon?
            # Wir zaehlen grob: inserted vs updated per separatem EXISTS-Check
            # waere teuer — hier begnuegen wir uns mit "touched".
            inserted += 1  # spaeter aufteilen, wenn wirklich noetig

    # Geloeschte Sources aus Inventory entfernen (sonst bleiben sie als
    # "Geister-Eintraege" und Flows wuerden sie wieder anfassen).
    # Wir matchen ueber den gebauten Inventory-rel_path:
    # "sources/<rel>" bzw. "context/<rel>" aus agent_sources (status='deleted').
    deleted_src_rows = conn.execute(
        f"""
        SELECT rel_path, kind FROM agent_sources
        WHERE status = 'deleted'
          AND kind IN ({placeholders})
          AND lower(extension) = 'pdf'
        """,
        tuple(kinds),
    ).fetchall()

    removed = 0
    for r in deleted_src_rows:
        scope_prefix = "sources" if r["kind"] == "source" else "context"
        inv_rel_path = f"{scope_prefix}/{r['rel_path']}"
        cur = conn.execute(
            "DELETE FROM agent_pdf_inventory WHERE rel_path = ?",
            (inv_rel_path,),
        )
        removed += cur.rowcount

    total_in_inventory = conn.execute(
        "SELECT COUNT(*) FROM agent_pdf_inventory"
    ).fetchone()[0]

    return {
        "inserted": inserted,
        "updated": updated,
        "removed": removed,
        "total_in_inventory": int(total_in_inventory),
    }


@register(
    name="sources_register",
    description=(
        "Scannt den gewaehlten Scope (sources/, context/ oder beides) rekursiv "
        "und aktualisiert die agent_sources-Registry. Erkennt neue, geaenderte "
        "und geloeschte Dateien ueber sha256-Hash-Vergleich. Idempotent — "
        "Wiederholung auf unveraendertem Stand liefert 0 Delta. Pfad-Parts "
        "mit '_'- oder '.'-Prefix (z.B. _meta/, _manifest.md, .DS_Store) "
        "gelten als intern und werden ignoriert. Nach dem Scan wird "
        "agent_pdf_inventory (Input der PDF-Pipeline) automatisch synchroni- "
        "siert — damit laufen Context-PDFs durch *dieselben* Flows "
        "(extraction_routing_decision, extraction) wie Source-PDFs, nur "
        "mit kind='context'."
    ),
    parameters={
        "type": "object",
        "properties": {
            "scope": {
                "type": "string",
                "enum": ["sources", "context", "both"],
                "default": "both",
                "description": (
                    "Welcher Unterbaum gescannt wird. "
                    "**Lass diesen Parameter in der Regel weg** — der "
                    "Default 'both' scannt sources/ UND context/ nacheinander, "
                    "was fast immer das Gewuenschte ist (damit context-PDFs "
                    "auch ins agent_pdf_inventory wandern und durch die "
                    "Pipeline laufen koennen). "
                    "Setze scope NUR explizit, wenn der Nutzer ausdruecklich "
                    "nur einen Unterbaum will: 'sources' nur sources/ "
                    "(kind='source'), 'context' nur context/ (kind='context')."
                ),
            },
            "subpath": {
                "type": "string",
                "description": (
                    "Optionaler Unterordner unter dem scope-Root "
                    "(z.B. 'Elektro' bei scope='sources'). "
                    "Leer = ganzer Baum. Bei scope='both' (Default) wirkt subpath "
                    "in beiden Baeumen gleich — in der Praxis meist leer lassen."
                ),
            },
            "skip_hash_if_unchanged": {
                "type": "boolean",
                "description": (
                    "True (Default): wenn (rel_path, size, mtime) gleich, "
                    "kein Re-Hash. False: immer rehashen (sicherer, langsamer)."
                ),
            },
            "scan_type": {
                "type": "string",
                "description": "Freies Label fuer die Scan-Historie, z.B. 'initial' oder 'nach-sp-export'.",
            },
        },
        "required": [],
    },
    returns=(
        "{scan_id, scan_type, scope, per_scope: {source: {...}, context: {...}}, "
        "stats: {new, changed, deleted, unchanged, total_active}, "
        "pdf_inventory_sync: {inserted, updated, removed}, dauer_s}"
    ),
)
def _sources_register(
    *,
    scope: str = "both",
    subpath: str = "",
    skip_hash_if_unchanged: bool = True,
    scan_type: str = "incremental",
) -> dict[str, Any]:
    import time

    if scope not in ("sources", "context", "both"):
        raise ValueError(
            f"scope muss 'sources', 'context' oder 'both' sein, nicht {scope!r}"
        )

    # Pipeline-Reform v2: in migrierten Projekten ist agent_sources hash-
    # zentriert und rel_path lebt in agent_source_locations. Wir
    # detektieren das Schema und dispatchen zu v1 oder v2 Scan-Funktionen.
    _check_conn = _connect_datastore_rw()
    try:
        is_v2 = _check_conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' "
            "AND name='agent_source_locations'"
        ).fetchone() is not None
    finally:
        _check_conn.close()

    root = _data_root()

    # Welche Scopes werden gescannt?
    scopes: list[tuple[str, str]]  # [(scope_name, kind), ...]
    if scope == "both":
        scopes = [("sources", "source"), ("context", "context")]
    else:
        kind = _SCOPE_TO_KIND[scope][0]
        scopes = [(scope, kind)]

    # Pre-Flight: scope-Roots existieren?
    for scope_name, _ in scopes:
        scope_root = (root / scope_name).resolve()
        if not scope_root.exists():
            if scope == "both":
                # Bei 'both' ist ein fehlender Scope tolerant — ueberspringen.
                continue
            raise ValueError(
                f"{scope_name}/ Ordner existiert nicht: {scope_root}. "
                "Lege Dateien dort ab oder fuehre 'disco project init' neu aus."
            )

    t_start = time.monotonic()
    conn = _connect_datastore_rw()

    # Einen gemeinsamen Scan-Eintrag anlegen
    cur = conn.execute(
        "INSERT INTO agent_source_scans (scan_type) VALUES (?)",
        (scan_type,),
    )
    scan_id = cur.lastrowid
    conn.commit()

    per_scope: dict[str, dict[str, Any]] = {}
    try:
        for scope_name, kind in scopes:
            scope_root = (root / scope_name).resolve()
            if not scope_root.exists():
                per_scope[kind] = {
                    "kind": kind,
                    "new": [], "changed": [], "deleted": [], "unchanged": 0,
                    "skipped_reason": f"{scope_name}/ nicht vorhanden",
                }
                continue
            scope_root.mkdir(parents=True, exist_ok=True)
            if is_v2:
                result = _scan_one_scope_v2(
                    conn=conn,
                    scan_id=scan_id,
                    kind=kind,
                    scope_root=scope_root,
                    subpath=subpath,
                    skip_hash_if_unchanged=skip_hash_if_unchanged,
                )
            else:
                result = _scan_one_scope(
                    conn=conn,
                    scan_id=scan_id,
                    kind=kind,
                    scope_root=scope_root,
                    subpath=subpath,
                    skip_hash_if_unchanged=skip_hash_if_unchanged,
                )
            per_scope[kind] = result

        # PDF-Inventory-Sync fuer alle angefassten kinds
        touched_kinds = [r.get("kind") for r in per_scope.values() if r.get("kind")]
        if is_v2:
            inventory_stats = _sync_pdf_inventory_v2(conn, touched_kinds)
        else:
            inventory_stats = _sync_pdf_inventory(conn, touched_kinds)

        # Aggregierte Statistik
        total_new = sum(len(r["new"]) for r in per_scope.values())
        total_changed = sum(len(r["changed"]) for r in per_scope.values())
        total_deleted = sum(len(r["deleted"]) for r in per_scope.values())
        total_unchanged = sum(r["unchanged"] for r in per_scope.values())

        elapsed = time.monotonic() - t_start
        total_active = conn.execute(
            "SELECT COUNT(*) FROM agent_sources WHERE status = 'active'"
        ).fetchone()[0]

        conn.execute(
            "UPDATE agent_source_scans SET "
            "  finished_at = datetime('now'), "
            "  n_new = ?, n_changed = ?, n_deleted = ?, n_unchanged = ? "
            "WHERE id = ?",
            (total_new, total_changed, total_deleted, total_unchanged, scan_id),
        )
        conn.commit()

    finally:
        conn.close()

    # Delta-Listen kuerzen fuer den Chat-Output (volle Listen sind in der DB)
    def _preview(lst: list[str], n: int = 20) -> dict:
        return {
            "count": len(lst),
            "sample": lst[:n],
            "truncated": len(lst) > n,
        }

    per_scope_preview: dict[str, Any] = {}
    for kind_name, r in per_scope.items():
        if "skipped_reason" in r:
            per_scope_preview[kind_name] = {"skipped_reason": r["skipped_reason"]}
            continue
        per_scope_preview[kind_name] = {
            "new": _preview(r["new"]),
            "changed": _preview(r["changed"]),
            "deleted": _preview(r["deleted"]),
            "unchanged": r["unchanged"],
        }

    return {
        "scan_id": scan_id,
        "scan_type": scan_type,
        "scope": scope,
        "scan_subpath": subpath or "(root)",
        "dauer_s": round(elapsed, 2),
        "per_scope": per_scope_preview,
        "stats": {
            "new": total_new,
            "changed": total_changed,
            "deleted": total_deleted,
            "unchanged": total_unchanged,
            "total_active_in_registry": total_active,
        },
        "pdf_inventory_sync": inventory_stats,
        "hint": _build_sources_hint(inventory_stats),
    }


def _build_sources_hint(inventory_stats: dict[str, Any]) -> str:
    """Baut den `hint`-String fuer `sources_register`-Return.

    Macht den naechsten Pipeline-Schritt explizit, wenn PDFs im Inventar
    stehen — Disco soll laut System-Prompt daraufhin aktiv die Pipeline
    anbieten, nicht generisch nachfragen.
    """
    base = (
        "Details pro Datei per SQL: "
        "SELECT rel_path, kind, size_bytes, extension, status, last_changed_at "
        "FROM ds.agent_sources ORDER BY last_changed_at DESC LIMIT 50. "
        "PDF-Inventory: SELECT rel_path, kind FROM ds.agent_pdf_inventory "
        "ORDER BY id DESC LIMIT 20."
    )
    total = int(inventory_stats.get("total_in_inventory", 0) or 0)
    if total > 0:
        return (
            base + " "
            f"PDF-Inventar: {total} Datei(en) bereit fuer die Pipeline "
            "(source + context zusammen). Naechster Schritt: "
            "`flow_run extraction_routing_decision` starten, danach "
            "`flow_run extraction` — dem Benutzer aktiv vorschlagen, "
            "nicht offen 'was moechtest Du als Naechstes' fragen."
        )
    return (
        base + " "
        "Keine PDFs im Paket — Standard-Pipeline (Routing/Extraktion) entfaellt."
    )


# ===========================================================================
# sources_attach_metadata — Begleit-Excel/CSV → agent_source_metadata
# ===========================================================================


def _normalize_rel_path(p: str) -> str:
    """Normalisiert User-Pfade (Backslashes, fuehrende ./, sources/ Prefix).
    Gibt den Pfad **relativ zu sources/** zurueck, so wie er in agent_sources.rel_path steht.
    """
    s = (p or "").strip().replace("\\", "/")
    # fuehrendes ./ weg
    while s.startswith("./"):
        s = s[2:]
    # sources/-Prefix weg, falls mitgegeben
    if s.startswith("sources/"):
        s = s[len("sources/"):]
    return s.strip("/")


def _load_metadata_rows(abs_path: Path, key_column: str, sheet: str | None) -> tuple[list[dict], list[str]]:
    """Liest eine xlsx/csv und gibt (rows_as_dicts, warnings) zurueck.

    rows: list of {col: value}, Spaltennamen wie in der Datei.
    """
    warnings: list[str] = []
    ext = abs_path.suffix.lower()

    if ext == ".csv":
        import csv, io
        text = abs_path.read_text(encoding="utf-8")
        if text.startswith("\ufeff"):
            text = text[1:]
        # Delimiter-Heuristik: erste Zeile ; vs ,
        first_line = text.split("\n", 1)[0]
        delim = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        rows = [dict(r) for r in reader]
        if not rows:
            warnings.append("CSV enthaelt keine Datenzeilen")
        return rows, warnings

    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(abs_path, data_only=True, read_only=True)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' nicht gefunden. Vorhanden: {wb.sheetnames}")
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header = [(h if h is not None else f"col_{i+1}") for i, h in enumerate(next(it))]
        except StopIteration:
            wb.close()
            raise ValueError("Sheet ist leer.")
        rows = []
        for row in it:
            if row is None:
                continue
            rec = {header[i]: ("" if row[i] is None else row[i]) for i in range(min(len(header), len(row)))}
            # komplett leere Zeilen ueberspringen
            if all((v == "" or v is None) for v in rec.values()):
                continue
            rows.append(rec)
        wb.close()
        return rows, warnings

    raise ValueError(f"Unbekannter Dateityp fuer Metadaten: {ext}")


def _match_row_to_source(
    row: dict, key_column: str, index_by_rel: dict[str, int],
    index_by_filename: dict[str, list[int]],
) -> tuple[int | None, str]:
    """Sucht fuer eine Zeile die passende agent_sources.id.

    Strategie (Option C mit A als Default):
      1. Exakt-Match auf rel_path (normalisiert)
      2. Fallback: Filename-Match (Basename ohne Pfad)
      3. Wenn Filename mehrdeutig (mehrere Treffer): ambig → None + Grund
      4. Wenn nichts gefunden: None + Grund

    Returns (source_id_or_None, status):
      status ∈ {'exact', 'filename', 'ambiguous', 'not-found', 'no-key'}
    """
    raw = row.get(key_column)
    if raw is None or str(raw).strip() == "":
        return None, "no-key"
    key = _normalize_rel_path(str(raw))
    # Schritt 1: exakt
    if key in index_by_rel:
        return index_by_rel[key], "exact"
    # Schritt 2: Fallback Filename (letztes Segment)
    fname = key.split("/")[-1]
    cands = index_by_filename.get(fname, [])
    if len(cands) == 1:
        return cands[0], "filename"
    if len(cands) > 1:
        return None, "ambiguous"
    return None, "not-found"


@register(
    name="sources_attach_metadata",
    description=(
        "Liest eine Begleit-Excel oder -CSV (typischerweise unter "
        "sources/_meta/) und ordnet die Zeilen den registrierten Quelldateien "
        "zu. Schreibt pro Zelle einen Eintrag in agent_source_metadata "
        "(source_of_truth='begleit-excel'). "
        "Matching in drei Stufen: (1) exakt auf rel_path, (2) fallback "
        "Filename, (3) bei Mehrdeutigkeit/Nicht-Gefunden als Report "
        "zurueckgeben ohne zu schreiben, sodass der Benutzer entscheiden kann. "
        "Idempotent: beim zweiten Lauf werden bestehende Werte ueberschrieben."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": (
                    "Pfad zur Begleit-Datei, relativ zum Projekt-Root. "
                    "Typisch: 'sources/_meta/sources-meta.xlsx'."
                ),
            },
            "key_column": {
                "type": "string",
                "description": (
                    "Spalte in der Begleit-Datei, die den Dateipfad enthaelt "
                    "(z.B. 'rel_path', 'Dateiname', 'Pfad'). Standard: 'rel_path'."
                ),
            },
            "sheet": {
                "type": "string",
                "description": "Optional: Sheet-Name bei xlsx. Default: erstes Sheet.",
            },
            "ignore_columns": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Spalten, die nicht als Metadaten gespeichert werden sollen "
                    "(z.B. die key_column selbst wird automatisch ausgeschlossen)."
                ),
            },
            "commit": {
                "type": "boolean",
                "description": (
                    "True = Metadaten schreiben. False = Trockenlauf, nur Report "
                    "liefern. Default: false — bei erstmaligem Lauf zuerst Report "
                    "anzeigen, dann mit commit=true bestaetigen."
                ),
            },
        },
        "required": ["path"],
    },
    returns=(
        "{path, key_column, sheet, total_rows, matched_exact, matched_filename, "
        "ambiguous: [{row_index, key, candidates}], not_found: [...], "
        "columns_written, rows_written, committed: bool}"
    ),
)
def _sources_attach_metadata(
    *,
    path: str,
    key_column: str = "rel_path",
    sheet: str | None = None,
    ignore_columns: list[str] | None = None,
    commit: bool = False,
) -> dict[str, Any]:
    root = _data_root()
    abs_path = (root / path).resolve()
    try:
        abs_path.relative_to(root)
    except ValueError:
        raise ValueError(f"Pfad ausserhalb des Projekts: {path!r}")
    if not abs_path.exists():
        raise ValueError(f"Datei nicht gefunden: {path!r}")

    ignored = set(ignore_columns or [])
    ignored.add(key_column)

    rows, warnings = _load_metadata_rows(abs_path, key_column, sheet)
    if not rows:
        return {
            "path": path,
            "key_column": key_column,
            "sheet": sheet,
            "total_rows": 0,
            "matched_exact": 0,
            "matched_filename": 0,
            "ambiguous": [],
            "not_found": [],
            "columns_written": [],
            "rows_written": 0,
            "committed": False,
            "warnings": warnings,
            "hint": "Die Begleit-Datei enthaelt keine Datenzeilen.",
        }

    # Pruefen dass key_column tatsaechlich im Header vorkommt
    header_cols = list(rows[0].keys())
    if key_column not in header_cols:
        raise ValueError(
            f"Spalte '{key_column}' nicht im Header gefunden. "
            f"Vorhandene Spalten: {header_cols}"
        )

    # Indexe aufbauen fuer schnelles Matching.
    # Pipeline-Reform v2 (2026-05-16): rel_path und filename leben jetzt in
    # agent_source_locations (n:1 zu agent_sources). Eine source kann mehrere
    # locations haben — fuers Matching ist jeder location-Pfad ein potentieller
    # Key, der auf dieselbe source_id verweist.
    conn = _connect_datastore_rw()
    try:
        is_v2 = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' "
            "AND name='agent_source_locations'"
        ).fetchone() is not None
        if is_v2:
            src_rows = conn.execute(
                "SELECT l.source_id AS id, l.rel_path, l.filename "
                "FROM agent_source_locations l "
                "JOIN agent_sources s ON s.id = l.source_id "
                "WHERE s.status = 'active' AND l.status = 'active'"
            ).fetchall()
        else:
            # Pre-Pipeline-v2-Fallback (alte Projekte, die noch nicht migriert sind)
            src_rows = conn.execute(
                "SELECT id, rel_path, filename FROM agent_sources WHERE status = 'active'"
            ).fetchall()
    finally:
        conn.close()
    # Achtung: bei mehreren locations pro source kann derselbe rel_path nur
    # einmal vorkommen (UNIQUE-Eigenschaft pro location), aber derselbe
    # filename ggf. mehrfach pro source — deshalb pro source dedup.
    index_by_rel: dict[str, int] = {r["rel_path"]: r["id"] for r in src_rows}
    index_by_filename: dict[str, list[int]] = {}
    for r in src_rows:
        sid = r["id"]
        bucket = index_by_filename.setdefault(r["filename"], [])
        if sid not in bucket:
            bucket.append(sid)

    # Zeilenweise matchen (erst Report bauen)
    matched_exact = 0
    matched_filename = 0
    ambiguous: list[dict] = []
    not_found: list[dict] = []
    resolved_rows: list[tuple[int, dict]] = []  # (source_id, row) fuer commit

    for i, row in enumerate(rows, start=1):
        source_id, status = _match_row_to_source(
            row, key_column, index_by_rel, index_by_filename,
        )
        if status == "exact":
            matched_exact += 1
            resolved_rows.append((source_id, row))
        elif status == "filename":
            matched_filename += 1
            resolved_rows.append((source_id, row))
        elif status == "ambiguous":
            raw = str(row.get(key_column) or "")
            cands = index_by_filename.get(_normalize_rel_path(raw).split("/")[-1], [])
            ambiguous.append({
                "row_index": i, "key": raw, "candidate_source_ids": cands,
            })
        elif status == "not-found":
            not_found.append({"row_index": i, "key": str(row.get(key_column) or "")})
        elif status == "no-key":
            not_found.append({"row_index": i, "key": "(leer)"})

    result: dict[str, Any] = {
        "path": path,
        "key_column": key_column,
        "sheet": sheet,
        "total_rows": len(rows),
        "matched_exact": matched_exact,
        "matched_filename": matched_filename,
        "ambiguous": ambiguous[:50],  # Preview, der Rest ist in der Datei
        "ambiguous_total": len(ambiguous),
        "not_found": not_found[:50],
        "not_found_total": len(not_found),
        "warnings": warnings,
    }

    if not commit:
        # Trockenlauf — nur Report
        result["committed"] = False
        result["columns_written"] = []
        result["rows_written"] = 0
        result["hint"] = (
            "Trockenlauf. Mit commit=true werden die Metadaten in "
            "agent_source_metadata geschrieben. Ambiguous/not-found bitte "
            "zuerst klaeren (Pfade in der Begleit-Excel korrigieren oder "
            "ignore-Liste erweitern)."
        )
        return result

    # Commit: schreiben
    metadata_cols = [c for c in header_cols if c not in ignored]
    if not metadata_cols:
        result["committed"] = False
        result["hint"] = "Keine Metadaten-Spalten zum Schreiben (alle ignoriert)."
        return result

    conn = _connect_datastore_rw()
    rows_written = 0
    try:
        for source_id, row in resolved_rows:
            for col in metadata_cols:
                val = row.get(col)
                if val is None or str(val).strip() == "":
                    continue
                # UPSERT per ON CONFLICT (Unique auf source_id+key+source_of_truth)
                conn.execute(
                    """
                    INSERT INTO agent_source_metadata
                        (source_id, key, value, source_of_truth, updated_at)
                    VALUES (?, ?, ?, 'begleit-excel', datetime('now'))
                    ON CONFLICT(source_id, key, source_of_truth) DO UPDATE SET
                        value = excluded.value,
                        updated_at = excluded.updated_at
                    """,
                    (source_id, col, str(val)),
                )
                rows_written += 1
        conn.commit()
    finally:
        conn.close()

    result["committed"] = True
    result["columns_written"] = metadata_cols
    result["rows_written"] = rows_written
    result["hint"] = (
        "Metadaten geschrieben. Abfrage mit: "
        "SELECT s.rel_path, m.key, m.value FROM agent_source_metadata m "
        "JOIN agent_sources s ON s.id = m.source_id WHERE m.source_of_truth='begleit-excel'"
    )
    return result


# ===========================================================================
# sources_detect_duplicates — sha256-Gruppen → duplicate-of Relations
# ===========================================================================


@register(
    name="sources_detect_duplicates",
    description=(
        "OBSOLET seit Pipeline-Reform v2 (2026-05-16). Im hash-zentrierten "
        "Datastore sind Duplikate strukturell: Dateien mit gleichem sha256 "
        "teilen sich eine agent_sources-Zeile und haben mehrere "
        "agent_source_locations-Zeilen. Eine separate Erkennung ist nicht "
        "mehr nötig. Dieses Tool gibt Top-Duplikat-Gruppen (Sources mit "
        ">1 Location) im selben Output-Format zurück, schreibt aber "
        "KEINE Relations mehr."
    ),
    parameters={
        "type": "object",
        "properties": {
            "min_group_size": {
                "type": "integer",
                "description": (
                    "Mindestgroesse einer Hash-Gruppe, um als Duplikat zu zaehlen. "
                    "Default 2 (jede Mehrfachkopie). 3+ liefert nur echte "
                    "'3-fach-oder-mehr'-Faelle."
                ),
            },
            "include_deleted": {
                "type": "boolean",
                "description": (
                    "Auch Dateien mit status='deleted' einbeziehen (Default false)."
                ),
            },
            "verbose": {
                "type": "boolean",
                "description": (
                    "Wenn true, werden bis zu 50 Duplikat-Gruppen mit jeweils max "
                    "10 Kopien pro Gruppe zurueckgegeben. Default false = nur "
                    "Summary + Top-5-Gruppen mit max 3 Kopien pro Gruppe."
                ),
            },
        },
        "required": [],
    },
    returns=(
        "{scanned, groups_found, new_relations, non_canonical_total, "
        "top_duplicate_sets: [{sha256, size_in_set, canonical, sample_copies}]}"
    ),
)
def _sources_detect_duplicates(
    *,
    min_group_size: int = 2,
    include_deleted: bool = False,
    verbose: bool = False,
) -> dict[str, Any]:
    """Hash-zentrierter Read-Only Stub (Pipeline-Reform v2).

    Zeigt agent_sources mit mehreren active Locations. Schreibt nichts.
    """
    min_group_size = max(2, int(min_group_size))
    loc_status_filter = "" if include_deleted else " AND l.status = 'active'"
    src_status_filter = "" if include_deleted else " AND s.status = 'active'"

    conn = _connect_datastore_rw()
    try:
        # Detektiere ob neues Modell aktiv ist
        has_locations = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' "
            "AND name='agent_source_locations'"
        ).fetchone() is not None
        if not has_locations:
            return {
                "error": (
                    "agent_source_locations existiert nicht. Projekt nicht "
                    "auf Pipeline-Reform v2 migriert."
                ),
            }

        # Top-Duplikate: Sources mit den meisten active Locations
        group_rows = conn.execute(
            f"""
            SELECT s.id, s.sha256,
                   COUNT(l.id) AS n_locations,
                   GROUP_CONCAT(l.rel_path, '\x1f') AS paths_concat
            FROM agent_sources s
            JOIN agent_source_locations l ON l.source_id = s.id
            WHERE 1=1 {src_status_filter} {loc_status_filter}
            GROUP BY s.id
            HAVING n_locations >= ?
            ORDER BY n_locations DESC, s.id
            """,
            (min_group_size,),
        ).fetchall()

        duplicate_sets: list[dict] = []
        for gr in group_rows:
            paths = (gr["paths_concat"] or "").split("\x1f")
            canonical_path = paths[0] if paths else None
            duplicate_sets.append({
                "sha256": gr["sha256"],
                "size_in_set": gr["n_locations"],
                "canonical": {"id": gr["id"], "rel_path": canonical_path},
                "copies": [{"rel_path": p} for p in paths[1:]],
            })

        scanned = conn.execute(
            f"SELECT COUNT(*) FROM agent_sources s WHERE 1=1{src_status_filter}"
        ).fetchone()[0]
        non_canonical_total = sum(s["size_in_set"] - 1 for s in duplicate_sets)
    finally:
        conn.close()

    max_groups = 50 if verbose else 5
    max_copies = 10 if verbose else 3
    compact_sets = []
    for s in duplicate_sets[:max_groups]:
        copies = s["copies"]
        compact_sets.append({
            "sha256": (s["sha256"] or "")[:16] + "...",
            "size_in_set": s["size_in_set"],
            "canonical": s["canonical"],
            "sample_copies": copies[:max_copies],
            "copies_total": len(copies),
            "copies_truncated": len(copies) > max_copies,
        })

    return {
        "scanned": scanned,
        "groups_found": len(duplicate_sets),
        "new_relations": 0,  # obsolet — keine Relations mehr
        "non_canonical_total": non_canonical_total,
        "top_duplicate_sets": compact_sets,
        "duplicate_sets_total": len(duplicate_sets),
        "truncated": len(duplicate_sets) > max_groups,
        "verbose": verbose,
        "hint": (
            "Im Hash-Modell sind Duplikate strukturell sichtbar. "
            "Vollständige Liste der Mehrfach-Locations: "
            "SELECT s.id, s.sha256, COUNT(l.id) AS n, "
            "GROUP_CONCAT(l.rel_path) AS paths "
            "FROM agent_sources s "
            "JOIN agent_source_locations l ON l.source_id = s.id "
            "WHERE l.status='active' GROUP BY s.id HAVING n > 1;"
        ),
    }
