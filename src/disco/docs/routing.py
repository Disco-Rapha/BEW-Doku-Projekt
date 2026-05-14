"""Generische Routing-Logik fuer den extraction_routing_decision-Flow.

Entscheidet pro Datei (egal welches Format) eine Engine. Heuristiken sind
pro file_kind organisiert; PDF nutzt die bewaehrte 3-Tier-Logik (Sticky-
Rules + PyMuPDF-Seitenanalyse).

Output: (engine, reason, heuristics_dict)
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from . import file_kind_from_path

logger = logging.getLogger(__name__)

ROUTER_VERSION = "router-v3.2"  # Optimierungen 14.05: path-hint + early-exit + width-precheck


# ---------------------------------------------------------------------------
# PDF-Seitenanalyse (PyMuPDF) + Page-Kind-Klassifikation
# ---------------------------------------------------------------------------
#
# Hier von disco.flows.library.pdf_routing_decision.runner reinkopiert,
# damit die Analyse-Logik nicht mehr von einem (zwischenzeitlich entfernten)
# Flow-Modul abhaengt. Nur PageStats / analyze_page / decide_kind werden
# verwendet — alle drei sind reine Funktionen, keine Flow-Lifecycle-Sachen.


@dataclass
class PageStats:
    chars: int
    n_paths: int
    text_coverage: float
    vector_coverage: float
    image_coverage: float
    n_images: int
    width_pt: float
    kind: str


def analyze_page(page) -> PageStats:
    """Ermittle Basis-Signale und Seiten-kind fuer eine Seite (PyMuPDF)."""
    rect = page.rect
    page_area = float(rect.width * rect.height) if rect.width and rect.height else 0.0

    text = page.get_text("text") or ""
    chars = len(text)

    text_area = 0.0
    try:
        blocks = page.get_text("blocks") or []
    except Exception:
        blocks = []
    for b in blocks:
        if len(b) >= 5 and isinstance(b[4], str) and b[4].strip():
            x0, y0, x1, y1 = b[0], b[1], b[2], b[3]
            text_area += max(0.0, float(x1 - x0) * float(y1 - y0))

    n_paths = 0
    vector_area = 0.0
    try:
        drawings = page.get_drawings() or []
    except Exception:
        drawings = []
    for d in drawings:
        items = d.get("items") or []
        n_paths += len(items)
        rect_d = d.get("rect")
        if rect_d is not None:
            vector_area += float(rect_d.width * rect_d.height)

    image_area = 0.0
    n_images = 0
    try:
        images = page.get_images(full=True) or []
    except Exception:
        images = []
    for img in images:
        xref = img[0]
        try:
            r = page.get_image_bbox(xref)
        except Exception:
            continue
        n_images += 1
        image_area += float(r.width * r.height)

    if page_area > 0:
        text_coverage = min(text_area / page_area, 1.0)
        vector_coverage = min(vector_area / page_area, 1.0)
        image_coverage = min(image_area / page_area, 1.0)
    else:
        text_coverage = vector_coverage = image_coverage = 0.0

    kind = decide_kind(
        chars=chars,
        n_paths=n_paths,
        text_coverage=text_coverage,
        vector_coverage=vector_coverage,
        image_coverage=image_coverage,
        n_images=n_images,
    )

    return PageStats(
        chars=chars,
        n_paths=n_paths,
        text_coverage=text_coverage,
        vector_coverage=vector_coverage,
        image_coverage=image_coverage,
        n_images=n_images,
        width_pt=float(rect.width) if rect.width else 0.0,
        kind=kind,
    )


def decide_kind(
    chars: int,
    n_paths: int,
    text_coverage: float,
    vector_coverage: float,
    image_coverage: float,
    n_images: int,
) -> str:
    """Seiten-kind-Hierarchie: empty | scan | vector-drawing | mixed | text."""
    # 1) empty
    if chars < 20 and n_paths == 0 and image_coverage < 0.10:
        return "empty"
    # 2) klassischer Scan
    if chars < 50 and image_coverage > 0.50:
        return "scan"
    # 3) Scan mit OCR-Layer
    if (
        (vector_coverage + image_coverage) > 0.50
        and text_coverage < 0.40
        and chars < 3500
    ):
        return "scan"
    # 4) Text-Dominanz
    if chars >= 1000 and text_coverage >= 0.30:
        has_heavy_graphics = (
            vector_coverage > 0.25
            or image_coverage > 0.15
            or n_paths > 50
            or n_images > 1
        )
        return "mixed" if has_heavy_graphics else "text"
    # 5) Vektorzeichnung
    if vector_coverage > 0.40 and chars < 500:
        return "vector-drawing"
    # 6) Rest
    if chars >= 100 and vector_coverage > 0.10:
        return "mixed"
    return "text"


# ---------------------------------------------------------------------------
# Top-Level: pro Datei eine Routing-Entscheidung
# ---------------------------------------------------------------------------


def decide(rel_path: str, abs_path: Path, file_role: str = "source") -> dict[str, Any]:
    """Routing-Entscheidung fuer eine Datei.

    Args:
      rel_path: Pfad relativ zum Projekt-Root (z.B. "sources/Geprueft/foo.pdf")
      abs_path: absoluter Filesystem-Pfad
      file_role: "source" oder "context" (aus agent_sources.kind)

    Returns: dict mit
      file_kind: 'pdf' | 'excel' | 'dwg' | 'image' | 'other'
      engine:    konkrete Engine-ID (z.B. 'pdf-azure-di-hr')
      reason:    Klartext-Begruendung
      heuristics: format-spezifische Heuristik-Werte
      router_version
    """
    file_kind = file_kind_from_path(abs_path)

    if file_kind == "pdf":
        # Optimierung C (14.05): Pfad-Heuristik vor teurer PyMuPDF-Analyse.
        # Bei klaren Plan-Pfaden (Plan/Plaene/Schaltplan/Zeichnung) +
        # vernuenftiger Mindestgroesse direkt auf pdf-azure-di-hr routen.
        # Das spart bei CAD-Plaenen die teure get_drawings()-Enumeration.
        hint = _path_hint_plan(rel_path, abs_path)
        if hint is not None:
            engine, reason, heur = hint
        else:
            engine, reason, heur = _decide_pdf(abs_path)
    elif file_kind == "excel":
        engine, reason, heur = _decide_excel(abs_path, file_role)
    elif file_kind == "dwg":
        engine, reason, heur = _decide_dwg(abs_path)
    elif file_kind == "image":
        engine, reason, heur = _decide_image(abs_path)
    else:
        engine, reason, heur = "skip", f"unsupported file_kind={file_kind}", {}

    return {
        "file_kind": file_kind,
        "engine": engine,
        "reason": reason,
        "heuristics": heur,
        "router_version": ROUTER_VERSION,
    }


# ---------------------------------------------------------------------------
# PDF-Routing — Sticky-Rules wie pdf_routing_decision-Runner
# ---------------------------------------------------------------------------

# Diese Konstanten spiegeln die heutige Pipeline (Stand 2026-04-25)
_PLAN_FORMAT_MIN_WIDTH_PT = 1000.0
_LARGE_IMAGE_MIN_COVERAGE = 0.60

# Optimierung C (14.05): Pfad-Substrings die auf Plan/Zeichnung hindeuten.
# Wenn der relative Pfad einen dieser Marker enthaelt UND die Datei
# vernuenftig gross ist (>=_PATH_HINT_MIN_BYTES), routen wir ohne PyMuPDF-
# Analyse direkt auf pdf-azure-di-hr. KKS-Labels in CAD-Exporten brauchen
# HR — Risiko falscher Routing ist minimal, Risiko 10x langsame
# get_drawings()-Enumeration auf Plan-Pdfs ist hoch.
_PATH_HINT_PLAN_TERMS = (
    "plan",          # deckt: Plan, Pläne, Schaltplan, Lageplan, Übersichtsplan
    "zeichnung",     # deckt: Zeichnung, Zeichnungen, Werkzeichnung
    "schemen",       # deckt: Schema, Schemen
    "isometr",       # deckt: Isometrie, Isometrien
)
_PATH_HINT_MIN_BYTES = 1_500_000  # 1.5 MB — echte CAD-Plaene sind fast immer >1.5MB.
                                  # Kleinere PDFs im Plan-Pfad sind oft Pruefberichte / Text-Doku
                                  # → durch PyMuPDF-Analyse genauer routen (vermeidet HR-Kosten).


def _path_hint_plan(rel_path: str, abs_path: Path) -> tuple[str, str, dict[str, Any]] | None:
    """Pfad+Groesse-basierter Shortcut: bei Plan-Hint → direkt HR.

    Returns: (engine, reason, heur) wenn Pfad-Hint greift, sonst None.
    """
    rel_lower = rel_path.lower()
    matched_term = next((t for t in _PATH_HINT_PLAN_TERMS if t in rel_lower), None)
    if not matched_term:
        return None
    try:
        size = abs_path.stat().st_size
    except OSError:
        return None
    if size < _PATH_HINT_MIN_BYTES:
        return None
    engine = "pdf-azure-di-hr"
    reason = (
        f"path-hint '{matched_term}' + size={size//1000} KB → vermutlich "
        f"Plan/Zeichnung, direkt HR ohne PyMuPDF-Analyse"
    )
    heur = {
        # Heuristik-Felder leer/None — wir haben kein PyMuPDF gelaufen.
        # Marker macht klar dass via path-hint geroutet wurde.
        "n_pages": None,
        "kind_counts": None,
        "n_scan_pages": None,
        "n_vdrawing_pages": None,
        "n_text_pages": None,
        "n_mixed_pages": None,
        "max_page_width_pt": None,
        "n_large_image_pages": None,
        "path_hint": matched_term,
        "size_bytes": size,
    }
    return engine, reason, heur


def _decide_pdf(abs_path: Path) -> tuple[str, str, dict[str, Any]]:
    """3-Tier-Routing aus PyMuPDF-Seitenanalyse mit Early-Exit.

    Optimierung A (14.05): Loop bricht ab, sobald die Routing-Entscheidung
    klar ist — bei langen CAD-Plaenen reichen oft 1-2 Seiten um HR zu
    triggern. Die heuristics-Felder werden mit den bis dahin gesammelten
    Werten gefuellt; 'analyzed_pages' markiert wieviele Seiten wirklich
    inspiziert wurden, 'analysis_complete' ob vollstaendig oder via
    Early-Exit beendet.
    """
    import fitz  # PyMuPDF

    doc = fitz.open(abs_path)
    n_pages = doc.page_count

    # WIDTH-PRECHECK (router-v3.2, 14.05): page.rect.width ist O(1) (PDF-Header-
    # Lookup, kein get_drawings). Wenn irgendeine Seite Plan-Format-Width hat,
    # ist die Engine-Entscheidung (pdf-azure-di-hr) bereits gefallen — wir muessen
    # gar nicht erst analyze_page() laufen lassen (das ist auf CAD-Plaenen der
    # eigentliche Killer, weil get_drawings() Tausende Vector-Paths enumeriert).
    # Beobachtung Prod 14.05: 1-Page-Plan-PDFs mit 2384pt-Width dauerten 160s
    # durch analyze_page(); mit Width-Precheck < 10ms.
    try:
        max_w_pre = 0.0
        for p in doc:
            w = float(p.rect.width) if p.rect.width else 0.0
            if w > max_w_pre:
                max_w_pre = w
            if w > _PLAN_FORMAT_MIN_WIDTH_PT:
                # Ein einziger Plan-Page-Treffer reicht — Engine ist klar.
                break
    except Exception:
        max_w_pre = 0.0

    if max_w_pre > _PLAN_FORMAT_MIN_WIDTH_PT:
        doc.close()
        return (
            "pdf-azure-di-hr",
            f"plan-format width-precheck max_w={max_w_pre:.0f}pt "
            f"(>{_PLAN_FORMAT_MIN_WIDTH_PT:.0f}) — analyze_page skipped",
            {
                "n_pages": n_pages,
                "kind_counts": None,
                "n_scan_pages": None,
                "n_vdrawing_pages": None,
                "n_text_pages": None,
                "n_mixed_pages": None,
                "max_page_width_pt": max_w_pre,
                "n_large_image_pages": None,
                "analyzed_pages": 0,
                "analysis_complete": False,
                "precheck_method": "width-only",
            },
        )

    stats: list[PageStats] = []
    early_exit = False
    early_exit_reason: str | None = None

    try:
        for i, page in enumerate(doc):
            ps = analyze_page(page)
            stats.append(ps)
            # Early-Exit: sobald 1 vector-drawing-Seite gefunden ist die
            # Engine-Wahl (pdf-azure-di-hr) gefallen. Restliche Seiten
            # aendern daran nichts.
            if ps.kind == "vector-drawing":
                early_exit = True
                early_exit_reason = "vector-drawing-Seite gefunden"
                break
            # Plan-Format ueber Seitenbreite: ebenfalls reicht 1 Treffer.
            if ps.width_pt > _PLAN_FORMAT_MIN_WIDTH_PT:
                early_exit = True
                early_exit_reason = "Plan-Format-Seite gefunden"
                break
            # Large-Image-Seite: auch HR.
            if ps.image_coverage > _LARGE_IMAGE_MIN_COVERAGE:
                early_exit = True
                early_exit_reason = "large-image-Seite gefunden"
                break
    finally:
        doc.close()

    n_analyzed = len(stats)
    kind_counts: dict[str, int] = {
        "empty": 0, "scan": 0, "vector-drawing": 0, "mixed": 0, "text": 0,
    }
    for ps in stats:
        kind_counts[ps.kind] = kind_counts.get(ps.kind, 0) + 1

    n_scan = kind_counts.get("scan", 0)
    n_vdraw = kind_counts.get("vector-drawing", 0)
    n_text = kind_counts.get("text", 0)
    n_mixed = kind_counts.get("mixed", 0)
    max_w = max((ps.width_pt for ps in stats), default=0.0)
    n_big_img = sum(1 for ps in stats if ps.image_coverage > _LARGE_IMAGE_MIN_COVERAGE)
    is_plan = max_w > _PLAN_FORMAT_MIN_WIDTH_PT

    # Sticky-Rules wie heute (Stand 2026-04-25, Bench-bestaetigt):
    if n_vdraw > 0:
        engine = "pdf-azure-di-hr"
        reason = f"{n_vdraw} vector-drawing-Seite(n) → KKS-Labels brauchen HR"
    elif is_plan:
        engine = "pdf-azure-di-hr"
        reason = f"plan-format max_w={max_w:.0f}pt > {_PLAN_FORMAT_MIN_WIDTH_PT:.0f}"
    elif n_big_img > 0:
        engine = "pdf-azure-di-hr"
        reason = f"{n_big_img} Seite(n) mit image_coverage > {_LARGE_IMAGE_MIN_COVERAGE:.2f}"
    elif n_scan > 0:
        engine = "pdf-azure-di"
        reason = f"{n_scan} A4-Scan-Seite(n) ohne Plan/Grossbild"
    else:
        engine = "pdf-azure-di"
        reason = f"text-dominant ({n_text}t/{n_mixed}m) → azure-di"

    if early_exit:
        reason = f"early-exit: {early_exit_reason} (after {n_analyzed}/{n_pages} pages)"

    heur = {
        "n_pages": n_pages,
        "kind_counts": kind_counts,
        "n_scan_pages": n_scan,
        "n_vdrawing_pages": n_vdraw,
        "n_text_pages": n_text,
        "n_mixed_pages": n_mixed,
        "max_page_width_pt": max_w,
        "n_large_image_pages": n_big_img,
        "analyzed_pages": n_analyzed,
        "analysis_complete": not early_exit,
    }
    return engine, reason, heur


# ---------------------------------------------------------------------------
# Excel-Routing
# ---------------------------------------------------------------------------


def _decide_excel(abs_path: Path, file_role: str) -> tuple[str, str, dict[str, Any]]:
    """Excel: immer Markdown-Extraktion via excel-openpyxl.

    Vor 2026-05-07 wurde context-Excel automatisch via excel-table-import
    auch als SQL-Tabellen importiert, was zu workspace.db-Bläh fuehrte
    (60+ context_*-Tabellen pro Projekt, viele ungenutzt). Default ist
    jetzt einheitlich Markdown — Disco kann per Search-Index drauf
    zugreifen, das reicht in 95% der Faelle. Wenn der User explizit
    Lookup-Tabellen fuer SQL-Joins braucht, ruft er `import_xlsx_to_table`
    als bewusste Aktion auf (Skill `excel-formatter`).

    Bestehende context_*-Tabellen in Prod-Projekten bleiben unveraendert
    (nicht-destruktiv) und koennen bei Gelegenheit manuell aufgeraeumt
    werden, sobald die zugehoerigen Excels Markdown haben.
    """
    engine = "excel-openpyxl"
    reason = f"{file_role}-Excel → Markdown-Extraktion (SQL-Import nur bewusst via import_xlsx_to_table)"

    # Quick-Inspect fuer heuristics_json
    heur: dict[str, Any] = {"file_role": file_role}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        heur["sheet_names"] = list(wb.sheetnames)
        heur["n_sheets"] = len(wb.sheetnames)
        wb.close()
    except Exception as exc:
        heur["inspect_error"] = str(exc)

    return engine, reason, heur


# ---------------------------------------------------------------------------
# DWG-Routing — heute eine Engine
# ---------------------------------------------------------------------------


def _decide_dwg(abs_path: Path) -> tuple[str, str, dict[str, Any]]:
    suffix = abs_path.suffix.lower()
    engine = "dwg-ezdxf-local"
    if suffix == ".dxf":
        reason = "DXF (Text-Format) → ezdxf direkt"
    else:
        reason = "DWG → ezdxf via ODA File Converter"
    return engine, reason, {"format": suffix.lstrip(".")}


# ---------------------------------------------------------------------------
# Image-Routing — heute eine Engine
# ---------------------------------------------------------------------------


def _decide_image(abs_path: Path) -> tuple[str, str, dict[str, Any]]:
    engine = "image-gpt5-vision"
    reason = "Bild → GPT-5.1 Vision (Beschreibung + OCR + strukturierte Erkennung)"

    heur: dict[str, Any] = {}
    try:
        from PIL import Image
        with Image.open(abs_path) as img:
            heur["width"] = img.size[0]
            heur["height"] = img.size[1]
            heur["mode"] = img.mode
    except Exception as exc:
        heur["inspect_error"] = str(exc)
    heur["size_bytes"] = abs_path.stat().st_size

    return engine, reason, heur


__all__ = ["decide", "ROUTER_VERSION"]
