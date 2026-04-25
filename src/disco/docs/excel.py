"""Excel-Extractor — openpyxl basierter Markdown-Renderer.

Engines:
  - excel-openpyxl       : Sheets als Markdown-Tabellen (Standard fuer sources/)
  - excel-table-import   : Wie excel-openpyxl PLUS direkt-Import in context_*-
                           Tabellen (via Flow, Side-Effect dort)

Output-Format:
  Pro Sheet ein Markdown-Block:
    ## Sheet: <SheetName> (<rows> Zeilen × <cols> Spalten)
    | Header1 | Header2 | ... |
    |---|---|---|
    | Wert | Wert | ... |
    ...

Strategie bei nicht-Tabellen-Sheets:
  - Wenn keine erkennbaren Header: Roh-Werte als Liste
  - Bei Merged Cells: Wert in der ersten Zelle, Rest leer
  - Formel-Zellen: aktueller berechneter Wert (data_only fallback)
  - Sehr breite Sheets (>30 Spalten): truncated mit Hinweis

Limitierungen:
  - Pivot-Tabellen sind kompliziert (User-Rueckfrage im Skill)
  - Bilder/Diagramme im Sheet werden ignoriert (nur Text-Inhalt)
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from . import EXTRACTION_PIPELINE_VERSION

logger = logging.getLogger(__name__)

_ENGINE_VERSIONS: dict[str, str] = {
    "excel-openpyxl": "1.0",
    "excel-table-import": "1.0",
}

# Hard-Limit fuer Markdown-Output pro Sheet (Zeilen)
_MAX_ROWS_PER_SHEET = 1000
# Hard-Limit fuer Spalten (alles dahinter wird mit "..." abgekuerzt)
_MAX_COLS_PER_SHEET = 30


def extract(path: Path, engine: str) -> tuple[str, dict[str, Any]]:
    if engine not in _ENGINE_VERSIONS:
        raise ValueError(f"Unbekannte Excel-Engine: {engine!r}")

    import openpyxl

    # data_only=True: zeigt berechnete Formel-Werte statt Formeln selbst
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    md_parts: list[str] = []
    unit_offsets: list[dict[str, Any]] = []
    cursor = 0
    sheet_metas: list[dict[str, Any]] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        rows: list[tuple] = []
        for r in rows_iter:
            # Trim leere Trailing-Zellen
            r2 = list(r)
            while r2 and (r2[-1] is None or r2[-1] == ""):
                r2.pop()
            if r2:  # nur nicht-leere Zeilen
                rows.append(tuple(r2))

        n_rows = len(rows)
        n_cols = max((len(r) for r in rows), default=0)
        n_cols_show = min(n_cols, _MAX_COLS_PER_SHEET)
        truncated_cols = n_cols > _MAX_COLS_PER_SHEET
        rows_show = rows[:_MAX_ROWS_PER_SHEET]
        truncated_rows = n_rows > _MAX_ROWS_PER_SHEET

        # --- Markdown rendern ---
        sheet_md_parts: list[str] = []
        sheet_md_parts.append(
            f"## Sheet: {_md_escape(sheet_name)} "
            f"({n_rows} Zeile{'n' if n_rows != 1 else ''} × {n_cols} Spalte{'n' if n_cols != 1 else ''})"
        )

        if not rows_show:
            sheet_md_parts.append("_(leer)_")
        else:
            # Erste Zeile als Header verwenden
            header = rows_show[0]
            header_cells = [
                _md_cell(header[i] if i < len(header) else "")
                for i in range(n_cols_show)
            ]
            if truncated_cols:
                header_cells.append("…")
            sheet_md_parts.append("| " + " | ".join(header_cells) + " |")
            sheet_md_parts.append("|" + "|".join(["---"] * len(header_cells)) + "|")

            # Daten-Zeilen
            for r in rows_show[1:]:
                cells = [
                    _md_cell(r[i] if i < len(r) else "")
                    for i in range(n_cols_show)
                ]
                if truncated_cols:
                    cells.append("…")
                sheet_md_parts.append("| " + " | ".join(cells) + " |")

            if truncated_rows:
                sheet_md_parts.append(
                    f"\n_(... {n_rows - _MAX_ROWS_PER_SHEET} weitere Zeilen "
                    f"abgekuerzt — Volldaten nur via SQL-Import)_"
                )

        sheet_md = "\n".join(sheet_md_parts) + "\n"

        # Offsets — vor dem Anhaengen Cursor merken
        char_start = cursor
        if md_parts:
            # Separator zwischen Sheets
            md_parts.append("\n")
            cursor += 1
            char_start = cursor
        md_parts.append(sheet_md)
        cursor += len(sheet_md)
        char_end = cursor

        unit_offsets.append({
            "unit_num": sheet_idx,
            "unit_label": sheet_name,
            "char_start": char_start,
            "char_end": char_end,
        })
        sheet_metas.append({
            "name": sheet_name,
            "n_rows": n_rows,
            "n_cols": n_cols,
            "truncated_rows": truncated_rows,
            "truncated_cols": truncated_cols,
        })

    wb.close()

    md = "".join(md_parts)
    char_count = len(md)
    n_units = len(unit_offsets)

    engine_version = _ENGINE_VERSIONS.get(engine, "1.0")
    meta: dict[str, Any] = {
        "file_kind": "excel",
        "engine": engine,
        "n_units": n_units,
        "char_count": char_count,
        "unit_offsets": unit_offsets,
        "estimated_cost_eur": 0.0,  # lokal, gratis
        "extractor_version": (
            f"{EXTRACTION_PIPELINE_VERSION}:{engine}:{engine_version}"
        ),
        "meta_json": {
            "n_sheets": n_units,
            "sheets": sheet_metas,
        },
    }
    return md, meta


def _md_escape(s: Any) -> str:
    """Escape fuer Markdown-Inhalte (Pipes, Linebreaks)."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("|", "\\|").replace("\n", " ").replace("\r", " ")
    return s.strip()


def _md_cell(v: Any) -> str:
    """Cell-Wert als Markdown-Cell-String."""
    if v is None:
        return ""
    if isinstance(v, float):
        # Vermeide '1.0' bei Ganzzahlen
        if v.is_integer():
            return str(int(v))
        return f"{v:g}"
    return _md_escape(v)


__all__ = ["extract"]
