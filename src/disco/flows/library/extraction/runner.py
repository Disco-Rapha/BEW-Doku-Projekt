"""Flow: extraction.

Generischer Extraktions-Flow fuer alle Formate (PDF/Excel/DWG/Bild).
Liest die Routing-Entscheidung aus work_extraction_routing, ruft den
passenden Extractor auf, schreibt das Ergebnis in agent_doc_markdown
und agent_doc_unit_offsets.

Side-Effect bei engine='excel-table-import':
  Pro Sheet zusaetzlich eine SQL-Tabelle unter context_<slug> in
  workspace.db. Damit sind Lookup-Daten direkt joinbar.

Provenance:
  Vor dem eigentlichen Markdown-Inhalt wird ein <!-- provenance -->-
  Block geschrieben. Enthaelt rel_path + folder + extracted_at +
  extractor_version. So findet der FTS-Index 'Geprueft' oder
  'Nicht_geprueft' aus der Ordner-Struktur.

Ebenen-Hinweis:
  - work_extraction_routing → Ebene 3, workspace.db (main, schreibbar)
  - ds.agent_sources         → Ebene 1, datastore.db (read-only)
  - ds.agent_doc_markdown    → Ebene 2, datastore.db
  - ds.agent_doc_unit_offsets → Ebene 2, datastore.db
  - context_<slug>            → Ebene 3, workspace.db (nur excel-table-import)
"""
from __future__ import annotations

import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from disco.docs import (
    EXTRACTION_PIPELINE_VERSION,
    all_known_engines,
    build_provenance_header,
    dispatch_extract,
)
from disco.flows.sdk import FlowRun, run_context


def load_items(
    run: FlowRun,
    *,
    limit: int | None,
    only_engine: str | None,
    only_kind: str | None,
    force_rerun: bool,
) -> List[Dict]:
    """Lade zu extrahierende Dateien aus work_extraction_routing + agent_sources.

    Default: Skip wenn unveraenderter source_hash schon in agent_doc_markdown.
    force_rerun=True ignoriert die Skip-Logik.
    """
    where = ["w.engine IS NOT NULL", "w.engine != ''", "w.engine != 'skip'"]
    params: list = []

    if only_engine:
        if only_engine not in all_known_engines():
            raise ValueError(
                f"only_engine={only_engine!r} unbekannt. "
                f"Erlaubt: {sorted(all_known_engines())}"
            )
        where.append("w.engine = ?")
        params.append(only_engine)

    if only_kind:
        where.append("w.file_kind = ?")
        params.append(only_kind)

    if not force_rerun:
        where.append(
            "NOT EXISTS ("
            "  SELECT 1 FROM ds.agent_doc_markdown m "
            "  WHERE m.file_id = w.file_id "
            "    AND m.source_hash IS NOT NULL "
            "    AND m.source_hash = s.sha256"
            ")"
        )

    sql = (
        "SELECT w.file_id, w.rel_path, w.engine, w.file_kind, "
        "       s.sha256 AS source_hash, s.kind AS file_role "
        "FROM work_extraction_routing w "
        "JOIN ds.agent_sources s ON s.id = w.file_id "
        f"WHERE {' AND '.join(where)} "
        "ORDER BY w.file_kind, w.file_id"
    )

    rows = run.db.query(sql, params)
    items: List[Dict] = list(rows)
    if limit is not None:
        items = items[:limit]

    run.log(
        f"Extraktions-Input: {len(items)} Dateien "
        f"(only_engine={only_engine!r}, only_kind={only_kind!r}, "
        f"force_rerun={force_rerun}, "
        f"limit={limit if limit is not None else 'none'})"
    )
    return items


def process_item(run: FlowRun, row: Dict) -> Dict:
    file_id = int(row["file_id"])
    rel_path = str(row["rel_path"])
    engine = str(row["engine"])
    file_kind = str(row.get("file_kind") or "")
    source_hash = row.get("source_hash")
    file_role = str(row.get("file_role") or "source")

    if engine not in all_known_engines():
        raise ValueError(
            f"file_id={file_id}: engine={engine!r} ungueltig. "
            f"Erwartet {sorted(all_known_engines())}."
        )

    abs_path = (run.project_root / rel_path).resolve()
    if not abs_path.is_file():
        raise FileNotFoundError(f"Datei nicht gefunden: {rel_path}")

    t0 = time.monotonic()
    md_body, meta = dispatch_extract(abs_path, engine)
    duration_ms = (time.monotonic() - t0) * 1000.0

    extracted_at = datetime.now(timezone.utc).isoformat()
    extractor_version = meta.get("extractor_version") or EXTRACTION_PIPELINE_VERSION

    # Provenance-Header voranstellen, dann unit_offsets entsprechend
    # nach hinten verschieben.
    provenance = build_provenance_header(
        file_id=file_id,
        rel_path=rel_path,
        file_kind=file_kind,
        engine=engine,
        extracted_at=extracted_at,
        extractor_version=extractor_version,
    )
    prov_len = len(provenance)
    md = provenance + md_body

    unit_offsets = meta.get("unit_offsets") or []
    shifted_offsets = [
        {
            "unit_num": u["unit_num"],
            "unit_label": u.get("unit_label") or f"u{u['unit_num']}",
            "char_start": u["char_start"] + prov_len,
            "char_end": u["char_end"] + prov_len,
        }
        for u in unit_offsets
    ]

    char_count = len(md)

    # --- Side-Effect: excel-table-import ---
    imported_tables: list[dict[str, Any]] = []
    if engine == "excel-table-import":
        try:
            imported_tables = _import_excel_to_context_tables(
                run=run,
                abs_path=abs_path,
                rel_path=rel_path,
            )
        except Exception as exc:
            run.log(f"[extract] WARN excel-table-import fehlgeschlagen: {exc}")
            # Markdown-Output bleibt erhalten, nur Tabellen-Import scheiterte

    meta_json_obj: dict[str, Any] = dict(meta.get("meta_json") or {})
    if imported_tables:
        meta_json_obj["imported_tables"] = imported_tables

    # In agent_doc_markdown schreiben
    run.db.insert_row(
        "ds.agent_doc_markdown",
        {
            "file_id": file_id,
            "rel_path": rel_path,
            "engine": engine,
            "file_kind": file_kind,
            "md_content": md,
            "char_count": char_count,
            "n_units": len(shifted_offsets),
            "meta_json": json.dumps(meta_json_obj, ensure_ascii=False),
            "source_hash": source_hash,
            "duration_ms": duration_ms,
            "run_id": run.run_id,
            "created_at": extracted_at,
            "extractor_version": extractor_version,
        },
        on_conflict="replace",
    )

    # Unit-Offsets (PDF: Seiten, Excel: Sheets, DWG: Sektionen, Bild: 1)
    run.db.execute(
        "DELETE FROM ds.agent_doc_unit_offsets WHERE file_id = ?",
        (file_id,),
    )
    if shifted_offsets:
        run.db.executemany(
            "INSERT INTO ds.agent_doc_unit_offsets "
            "(file_id, unit_num, unit_label, char_start, char_end) "
            "VALUES (?, ?, ?, ?, ?)",
            [
                (
                    file_id,
                    u["unit_num"],
                    u["unit_label"],
                    u["char_start"],
                    u["char_end"],
                )
                for u in shifted_offsets
            ],
        )

    cost = float(meta.get("estimated_cost_eur", 0.0))
    if cost > 0:
        run.add_cost(eur=cost)

    run.log(
        f"[extract] file_id={file_id}, kind={file_kind}, engine={engine}, "
        f"units={len(shifted_offsets)}, chars={char_count}, "
        f"duration={duration_ms:.0f}ms, cost={cost:.4f} EUR"
        + (f", tables={[t['table'] for t in imported_tables]}" if imported_tables else "")
    )

    return {
        "file_id": file_id,
        "rel_path": rel_path,
        "engine": engine,
        "file_kind": file_kind,
        "n_units": len(shifted_offsets),
        "char_count": char_count,
        "duration_ms": duration_ms,
        "estimated_cost_eur": cost,
        "imported_tables": [t["table"] for t in imported_tables],
    }


# ---------------------------------------------------------------------------
# Excel → context_*-Tabellen importieren
# ---------------------------------------------------------------------------


def _slugify(s: str) -> str:
    """Tabellen-tauglicher Slug: nur a-z0-9_, max 50 Zeichen."""
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    s = re.sub(r"_+", "_", s)
    return s[:50] or "sheet"


def _import_excel_to_context_tables(
    *,
    run: FlowRun,
    abs_path: Path,
    rel_path: str,
) -> list[dict[str, Any]]:
    """Pro Sheet eine context_*-Tabelle in workspace.db anlegen.

    Tabellen-Naming:
      context_<file_slug>            (bei nur einem Sheet)
      context_<file_slug>__<sheet_slug>  (bei mehreren Sheets)

    Datentyp ist immer TEXT (einfach, robust). Der Agent kann gezielt
    casten oder eine spezifischere Importtabelle anlegen, wenn noetig.
    """
    import openpyxl

    file_slug = _slugify(Path(rel_path).stem)
    wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)

    imported: list[dict[str, Any]] = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        rows: list[tuple] = []
        for r in rows_iter:
            r2 = list(r)
            while r2 and (r2[-1] is None or r2[-1] == ""):
                r2.pop()
            if r2:
                rows.append(tuple(r2))

        if not rows:
            continue

        # Heuristik: erste Zeile = Header. Wenn nur 1 Zeile insgesamt:
        # skippen (kein Sinn fuer Tabelle).
        if len(rows) < 2:
            continue

        header = rows[0]
        n_cols = len(header)
        # Spaltennamen normalisieren
        col_names = [
            f"col_{_slugify(str(h or f'c{i}'))}"
            for i, h in enumerate(header, start=1)
        ]
        # Eindeutigkeit erzwingen
        seen: dict[str, int] = {}
        for i, name in enumerate(col_names):
            if name in seen:
                seen[name] += 1
                col_names[i] = f"{name}_{seen[name]}"
            else:
                seen[name] = 0

        if len(sheet_names) == 1:
            table_name = f"context_{file_slug}"
        else:
            table_name = f"context_{file_slug}__{_slugify(sheet_name)}"

        # Vorhandene Tabelle droppen (idempotente Re-Import-Logik)
        run.db.execute(f"DROP TABLE IF EXISTS {table_name}")

        cols_ddl = ",\n  ".join(f'"{c}" TEXT' for c in col_names)
        run.db.execute(f"CREATE TABLE {table_name} (\n  {cols_ddl}\n)")

        placeholders = ", ".join("?" * n_cols)
        insert_sql = (
            f"INSERT INTO {table_name} "
            f"({', '.join(chr(34) + c + chr(34) for c in col_names)}) "
            f"VALUES ({placeholders})"
        )

        # Daten einfügen
        data_rows = []
        for r in rows[1:]:
            vals = []
            for i in range(n_cols):
                v = r[i] if i < len(r) else None
                vals.append("" if v is None else str(v))
            data_rows.append(tuple(vals))

        if data_rows:
            run.db.executemany(insert_sql, data_rows)

        imported.append({
            "table": table_name,
            "sheet": sheet_name,
            "n_rows": len(data_rows),
            "n_cols": n_cols,
            "columns": col_names,
        })

    wb.close()
    return imported


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "ja"}
    return False


def main() -> None:
    with run_context(FlowRun.from_env()) as run:
        run.log(
            f"Flow {run.flow_name} gestartet "
            f"(run_id={run.run_id}, pipeline={EXTRACTION_PIPELINE_VERSION})"
        )

        cfg = run.config or {}

        limit = cfg.get("limit")
        if isinstance(limit, str):
            try:
                limit = int(limit)
            except ValueError:
                limit = None
        elif isinstance(limit, (int, float)):
            limit = int(limit)
        else:
            limit = None

        only_engine = cfg.get("only_engine")
        if only_engine is not None and not isinstance(only_engine, str):
            only_engine = str(only_engine)
        if only_engine == "":
            only_engine = None

        only_kind = cfg.get("only_kind")
        if only_kind is not None and not isinstance(only_kind, str):
            only_kind = str(only_kind)
        if only_kind == "":
            only_kind = None

        force_rerun = _parse_bool(cfg.get("force_rerun"))

        items = load_items(
            run,
            limit=limit,
            only_engine=only_engine,
            only_kind=only_kind,
            force_rerun=force_rerun,
        )
        run.set_total(len(items))

        if not items:
            run.log("Keine offenen Dateien fuer Extraktion – nichts zu tun.")
            return

        for row in items:
            file_id = row["file_id"]

            def work(it=row) -> Dict:
                return process_item(run, it)

            run.process(
                input_ref=f"file:{file_id}",
                fn=work,
                max_retries=1,
            )


if __name__ == "__main__":
    main()
