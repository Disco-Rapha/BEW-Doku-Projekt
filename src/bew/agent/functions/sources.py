"""Sources-Registry: rekursiver Scan von sources/ + Begleit-Metadaten + Duplikat-Erkennung.

Pattern: "File Registry with Change Detection"
  - sha256-basiert fuer sichere Aenderungserkennung
  - Idempotent: wiederholter Lauf auf unveraendertem Filesystem ergibt 0 Delta
  - Status-Feld 'active'|'deleted' fuer Abgang-Erkennung

Performance: auf SSD ca. 2-5s fuer 1600 Dateien bei gemischten Groessen.
Hash wird gecacht per (rel_path, size, mtime) — wenn nichts davon sich
geaendert hat, ueberspringen wir das erneute Hashen (optimierung).
"""

from __future__ import annotations

import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any

from . import register
from .data import _connect as db_connect
from .fs import _data_root


# Ordner/Namen die niemals gescannt werden (Metadaten-Container)
SCAN_IGNORE_TOP: frozenset[str] = frozenset({"_meta", ".DS_Store"})
SCAN_IGNORE_SUFFIX: tuple[str, ...] = (".tmp", ".crdownload", ".part")
SCAN_IGNORE_PATTERNS: tuple[str, ...] = (".git", "__pycache__", "node_modules")


def _sha256_file(path: Path, chunk_size: int = 1 << 20) -> str:
    """SHA-256 ueber den Datei-Inhalt, chunk-basiert (Memory-safe bei grossen PDFs)."""
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _iso_mtime(epoch: float) -> str:
    return datetime.fromtimestamp(epoch).strftime("%Y-%m-%d %H:%M:%S")


def _is_ignored(rel_path: Path) -> bool:
    """True wenn dieser Pfad beim Scan uebersprungen wird."""
    parts = rel_path.parts
    # Top-Level-Ignore (z.B. _meta/)
    if parts and parts[0] in SCAN_IGNORE_TOP:
        return True
    # Sonder-Pattern irgendwo im Pfad
    for p in parts:
        if p in SCAN_IGNORE_PATTERNS:
            return True
    # Suffix-Pattern
    for suf in SCAN_IGNORE_SUFFIX:
        if rel_path.name.endswith(suf):
            return True
    return False


@register(
    name="sources_register",
    description=(
        "Scannt den sources/-Ordner rekursiv und aktualisiert die "
        "agent_sources-Registry. Erkennt neue, geaenderte und geloeschte Dateien "
        "ueber sha256-Hash-Vergleich. Idempotent — Wiederholung auf unveraendertem "
        "Stand liefert 0 Delta. Ordner '_meta' wird ignoriert (dort gehoeren "
        "Begleit-Metadaten hin, nicht zu scannende Dateien). "
        "Gibt Statistik + Liste der Delta-Dateien zurueck."
    ),
    parameters={
        "type": "object",
        "properties": {
            "subpath": {
                "type": "string",
                "description": (
                    "Optionaler Unterordner unter sources/ (z.B. 'Elektro'), "
                    "wenn nicht alles neu gescannt werden soll. "
                    "Leer = ganzer sources/-Baum."
                ),
            },
            "skip_hash_if_unchanged": {
                "type": "boolean",
                "description": (
                    "True (Default): wenn (rel_path, size, mtime) gleich, "
                    "kein Re-Hash. False: immer rehashen (sicherer, langsamer)."
                ),
            },
            "scan_type": {
                "type": "string",
                "description": "Freies Label fuer die Scan-Historie, z.B. 'initial' oder 'nach-sp-export'.",
            },
        },
        "required": [],
    },
    returns=(
        "{scan_id, scan_type, stats: {new, changed, deleted, unchanged, total_active}, "
        "delta: {new: [...], changed: [...], deleted: [...]}, dauer_s}"
    ),
)
def _sources_register(
    *,
    subpath: str = "",
    skip_hash_if_unchanged: bool = True,
    scan_type: str = "incremental",
) -> dict[str, Any]:
    import time

    root = _data_root()
    sources_root = (root / "sources").resolve()
    if not sources_root.exists():
        raise ValueError(
            "sources/ Ordner existiert nicht. Lege Quelldateien unter "
            f"{sources_root} ab oder fuehre 'disco project init' erneut aus."
        )
    sources_root.mkdir(parents=True, exist_ok=True)

    scan_subdir = (sources_root / subpath).resolve() if subpath else sources_root
    try:
        scan_subdir.relative_to(sources_root)
    except ValueError:
        raise ValueError(f"subpath ausserhalb von sources/: {subpath!r}")

    t_start = time.monotonic()
    conn = db_connect()

    # Scan-Eintrag anlegen
    cur = conn.execute(
        "INSERT INTO agent_source_scans (scan_type) VALUES (?)",
        (scan_type,),
    )
    scan_id = cur.lastrowid
    conn.commit()

    try:
        # Bestehende aktive Eintraege laden (fuer Diff)
        # Key = rel_path, Value = dict mit bisherigen Werten
        existing: dict[str, dict] = {}
        # Wenn subpath gegeben: nur die Sub-Eintraege laden
        if subpath:
            like = f"{subpath.rstrip('/')}/%"
            rows = conn.execute(
                "SELECT id, rel_path, size_bytes, sha256, mtime, status "
                "FROM agent_sources WHERE rel_path LIKE ? OR rel_path = ?",
                (like, subpath.rstrip("/")),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, rel_path, size_bytes, sha256, mtime, status "
                "FROM agent_sources"
            ).fetchall()
        for r in rows:
            existing[r["rel_path"]] = dict(r)

        seen_rel_paths: set[str] = set()
        new_list: list[str] = []
        changed_list: list[str] = []
        unchanged_count = 0

        # Filesystem abgehen
        for fs_path in scan_subdir.rglob("*"):
            if not fs_path.is_file():
                continue
            rel = fs_path.relative_to(sources_root)
            if _is_ignored(rel):
                continue
            rel_str = str(rel)
            seen_rel_paths.add(rel_str)

            try:
                st = fs_path.stat()
            except OSError:
                continue
            size = st.st_size
            mtime_iso = _iso_mtime(st.st_mtime)
            folder = str(rel.parent) if rel.parent != Path(".") else ""
            ext = rel.suffix.lstrip(".").lower() or None

            prev = existing.get(rel_str)

            # Hashing: nur wenn unbekannt oder sich (size, mtime) geaendert hat
            need_hash = True
            if prev and skip_hash_if_unchanged:
                if prev["size_bytes"] == size and prev["mtime"] == mtime_iso:
                    need_hash = False

            digest = None
            if need_hash:
                try:
                    digest = _sha256_file(fs_path)
                except OSError as exc:
                    # Datei nicht lesbar — als Warning, aber weitermachen
                    conn.execute(
                        "UPDATE agent_source_scans SET notes = "
                        "COALESCE(notes, '') || ? WHERE id = ?",
                        (f"Lesefehler {rel_str}: {exc}\n", scan_id),
                    )
                    continue
            else:
                digest = prev["sha256"]

            if not prev:
                # NEU
                conn.execute(
                    "INSERT INTO agent_sources "
                    "(rel_path, filename, folder, extension, size_bytes, sha256, "
                    " mtime, first_seen_at, last_seen_at, last_changed_at, status) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'), "
                    "        datetime('now'), 'active')",
                    (rel_str, fs_path.name, folder, ext, size, digest, mtime_iso),
                )
                new_list.append(rel_str)
            else:
                # Bekannt — Hash-Diff bestimmt "changed"
                hash_changed = (prev["sha256"] or "") != (digest or "")
                was_deleted = prev["status"] == "deleted"
                if hash_changed:
                    conn.execute(
                        "UPDATE agent_sources SET "
                        "  size_bytes = ?, sha256 = ?, mtime = ?, "
                        "  last_seen_at = datetime('now'), "
                        "  last_changed_at = datetime('now'), "
                        "  status = 'active', "
                        "  updated_at = datetime('now') "
                        "WHERE id = ?",
                        (size, digest, mtime_iso, prev["id"]),
                    )
                    changed_list.append(rel_str)
                else:
                    # Unveraendert — nur last_seen_at pflegen
                    # (und ggf. status 'deleted' → 'active' wenn Datei wieder auftaucht)
                    if was_deleted:
                        conn.execute(
                            "UPDATE agent_sources SET "
                            "  status = 'active', last_seen_at = datetime('now'), "
                            "  last_changed_at = datetime('now'), "
                            "  updated_at = datetime('now') "
                            "WHERE id = ?",
                            (prev["id"],),
                        )
                        changed_list.append(rel_str)  # zaehlt auch als "changed" (reaktiviert)
                    else:
                        conn.execute(
                            "UPDATE agent_sources SET "
                            "  last_seen_at = datetime('now') "
                            "WHERE id = ?",
                            (prev["id"],),
                        )
                        unchanged_count += 1

        # Dateien, die in DB aber nicht mehr im FS → status='deleted'
        deleted_list: list[str] = []
        for rel_str, prev in existing.items():
            if rel_str not in seen_rel_paths and prev["status"] == "active":
                conn.execute(
                    "UPDATE agent_sources SET "
                    "  status = 'deleted', "
                    "  last_changed_at = datetime('now'), "
                    "  updated_at = datetime('now') "
                    "WHERE id = ?",
                    (prev["id"],),
                )
                deleted_list.append(rel_str)

        # Scan-Eintrag finalisieren
        elapsed = time.monotonic() - t_start
        total_active = conn.execute(
            "SELECT COUNT(*) FROM agent_sources WHERE status = 'active'"
        ).fetchone()[0]
        conn.execute(
            "UPDATE agent_source_scans SET "
            "  finished_at = datetime('now'), "
            "  n_new = ?, n_changed = ?, n_deleted = ?, n_unchanged = ? "
            "WHERE id = ?",
            (len(new_list), len(changed_list), len(deleted_list),
             unchanged_count, scan_id),
        )
        conn.commit()

    finally:
        conn.close()

    # Delta-Listen kuerzen fuer den Chat-Output (volle Listen sind in der DB)
    def _preview(lst: list[str], n: int = 20) -> dict:
        return {
            "count": len(lst),
            "sample": lst[:n],
            "truncated": len(lst) > n,
        }

    return {
        "scan_id": scan_id,
        "scan_type": scan_type,
        "scan_subpath": subpath or "(root)",
        "dauer_s": round(elapsed, 2),
        "stats": {
            "new": len(new_list),
            "changed": len(changed_list),
            "deleted": len(deleted_list),
            "unchanged": unchanged_count,
            "total_active_in_registry": total_active,
        },
        "delta": {
            "new": _preview(new_list),
            "changed": _preview(changed_list),
            "deleted": _preview(deleted_list),
        },
        "hint": (
            "Details pro Datei per SQL: "
            "SELECT rel_path, size_bytes, extension, status, last_changed_at "
            "FROM agent_sources ORDER BY last_changed_at DESC LIMIT 50"
        ),
    }


# ===========================================================================
# sources_attach_metadata — Begleit-Excel/CSV → agent_source_metadata
# ===========================================================================


def _normalize_rel_path(p: str) -> str:
    """Normalisiert User-Pfade (Backslashes, fuehrende ./, sources/ Prefix).
    Gibt den Pfad **relativ zu sources/** zurueck, so wie er in agent_sources.rel_path steht.
    """
    s = (p or "").strip().replace("\\", "/")
    # fuehrendes ./ weg
    while s.startswith("./"):
        s = s[2:]
    # sources/-Prefix weg, falls mitgegeben
    if s.startswith("sources/"):
        s = s[len("sources/"):]
    return s.strip("/")


def _load_metadata_rows(abs_path: Path, key_column: str, sheet: str | None) -> tuple[list[dict], list[str]]:
    """Liest eine xlsx/csv und gibt (rows_as_dicts, warnings) zurueck.

    rows: list of {col: value}, Spaltennamen wie in der Datei.
    """
    warnings: list[str] = []
    ext = abs_path.suffix.lower()

    if ext == ".csv":
        import csv, io
        text = abs_path.read_text(encoding="utf-8")
        if text.startswith("\ufeff"):
            text = text[1:]
        # Delimiter-Heuristik: erste Zeile ; vs ,
        first_line = text.split("\n", 1)[0]
        delim = ";" if first_line.count(";") > first_line.count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        rows = [dict(r) for r in reader]
        if not rows:
            warnings.append("CSV enthaelt keine Datenzeilen")
        return rows, warnings

    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(abs_path, data_only=True, read_only=True)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' nicht gefunden. Vorhanden: {wb.sheetnames}")
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header = [(h if h is not None else f"col_{i+1}") for i, h in enumerate(next(it))]
        except StopIteration:
            wb.close()
            raise ValueError("Sheet ist leer.")
        rows = []
        for row in it:
            if row is None:
                continue
            rec = {header[i]: ("" if row[i] is None else row[i]) for i in range(min(len(header), len(row)))}
            # komplett leere Zeilen ueberspringen
            if all((v == "" or v is None) for v in rec.values()):
                continue
            rows.append(rec)
        wb.close()
        return rows, warnings

    raise ValueError(f"Unbekannter Dateityp fuer Metadaten: {ext}")


def _match_row_to_source(
    row: dict, key_column: str, index_by_rel: dict[str, int],
    index_by_filename: dict[str, list[int]],
) -> tuple[int | None, str]:
    """Sucht fuer eine Zeile die passende agent_sources.id.

    Strategie (Option C mit A als Default):
      1. Exakt-Match auf rel_path (normalisiert)
      2. Fallback: Filename-Match (Basename ohne Pfad)
      3. Wenn Filename mehrdeutig (mehrere Treffer): ambig → None + Grund
      4. Wenn nichts gefunden: None + Grund

    Returns (source_id_or_None, status):
      status ∈ {'exact', 'filename', 'ambiguous', 'not-found', 'no-key'}
    """
    raw = row.get(key_column)
    if raw is None or str(raw).strip() == "":
        return None, "no-key"
    key = _normalize_rel_path(str(raw))
    # Schritt 1: exakt
    if key in index_by_rel:
        return index_by_rel[key], "exact"
    # Schritt 2: Fallback Filename (letztes Segment)
    fname = key.split("/")[-1]
    cands = index_by_filename.get(fname, [])
    if len(cands) == 1:
        return cands[0], "filename"
    if len(cands) > 1:
        return None, "ambiguous"
    return None, "not-found"


@register(
    name="sources_attach_metadata",
    description=(
        "Liest eine Begleit-Excel oder -CSV (typischerweise unter "
        "sources/_meta/) und ordnet die Zeilen den registrierten Quelldateien "
        "zu. Schreibt pro Zelle einen Eintrag in agent_source_metadata "
        "(source_of_truth='begleit-excel'). "
        "Matching in drei Stufen: (1) exakt auf rel_path, (2) fallback "
        "Filename, (3) bei Mehrdeutigkeit/Nicht-Gefunden als Report "
        "zurueckgeben ohne zu schreiben, sodass der Benutzer entscheiden kann. "
        "Idempotent: beim zweiten Lauf werden bestehende Werte ueberschrieben."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": (
                    "Pfad zur Begleit-Datei, relativ zum Projekt-Root. "
                    "Typisch: 'sources/_meta/sources-meta.xlsx'."
                ),
            },
            "key_column": {
                "type": "string",
                "description": (
                    "Spalte in der Begleit-Datei, die den Dateipfad enthaelt "
                    "(z.B. 'rel_path', 'Dateiname', 'Pfad'). Standard: 'rel_path'."
                ),
            },
            "sheet": {
                "type": "string",
                "description": "Optional: Sheet-Name bei xlsx. Default: erstes Sheet.",
            },
            "ignore_columns": {
                "type": "array",
                "items": {"type": "string"},
                "description": (
                    "Spalten, die nicht als Metadaten gespeichert werden sollen "
                    "(z.B. die key_column selbst wird automatisch ausgeschlossen)."
                ),
            },
            "commit": {
                "type": "boolean",
                "description": (
                    "True = Metadaten schreiben. False = Trockenlauf, nur Report "
                    "liefern. Default: false — bei erstmaligem Lauf zuerst Report "
                    "anzeigen, dann mit commit=true bestaetigen."
                ),
            },
        },
        "required": ["path"],
    },
    returns=(
        "{path, key_column, sheet, total_rows, matched_exact, matched_filename, "
        "ambiguous: [{row_index, key, candidates}], not_found: [...], "
        "columns_written, rows_written, committed: bool}"
    ),
)
def _sources_attach_metadata(
    *,
    path: str,
    key_column: str = "rel_path",
    sheet: str | None = None,
    ignore_columns: list[str] | None = None,
    commit: bool = False,
) -> dict[str, Any]:
    root = _data_root()
    abs_path = (root / path).resolve()
    try:
        abs_path.relative_to(root)
    except ValueError:
        raise ValueError(f"Pfad ausserhalb des Projekts: {path!r}")
    if not abs_path.exists():
        raise ValueError(f"Datei nicht gefunden: {path!r}")

    ignored = set(ignore_columns or [])
    ignored.add(key_column)

    rows, warnings = _load_metadata_rows(abs_path, key_column, sheet)
    if not rows:
        return {
            "path": path,
            "key_column": key_column,
            "sheet": sheet,
            "total_rows": 0,
            "matched_exact": 0,
            "matched_filename": 0,
            "ambiguous": [],
            "not_found": [],
            "columns_written": [],
            "rows_written": 0,
            "committed": False,
            "warnings": warnings,
            "hint": "Die Begleit-Datei enthaelt keine Datenzeilen.",
        }

    # Pruefen dass key_column tatsaechlich im Header vorkommt
    header_cols = list(rows[0].keys())
    if key_column not in header_cols:
        raise ValueError(
            f"Spalte '{key_column}' nicht im Header gefunden. "
            f"Vorhandene Spalten: {header_cols}"
        )

    # agent_sources-Indexe aufbauen fuer schnelles Matching
    conn = db_connect()
    try:
        src_rows = conn.execute(
            "SELECT id, rel_path, filename FROM agent_sources WHERE status = 'active'"
        ).fetchall()
    finally:
        conn.close()
    index_by_rel: dict[str, int] = {r["rel_path"]: r["id"] for r in src_rows}
    index_by_filename: dict[str, list[int]] = {}
    for r in src_rows:
        index_by_filename.setdefault(r["filename"], []).append(r["id"])

    # Zeilenweise matchen (erst Report bauen)
    matched_exact = 0
    matched_filename = 0
    ambiguous: list[dict] = []
    not_found: list[dict] = []
    resolved_rows: list[tuple[int, dict]] = []  # (source_id, row) fuer commit

    for i, row in enumerate(rows, start=1):
        source_id, status = _match_row_to_source(
            row, key_column, index_by_rel, index_by_filename,
        )
        if status == "exact":
            matched_exact += 1
            resolved_rows.append((source_id, row))
        elif status == "filename":
            matched_filename += 1
            resolved_rows.append((source_id, row))
        elif status == "ambiguous":
            raw = str(row.get(key_column) or "")
            cands = index_by_filename.get(_normalize_rel_path(raw).split("/")[-1], [])
            ambiguous.append({
                "row_index": i, "key": raw, "candidate_source_ids": cands,
            })
        elif status == "not-found":
            not_found.append({"row_index": i, "key": str(row.get(key_column) or "")})
        elif status == "no-key":
            not_found.append({"row_index": i, "key": "(leer)"})

    result: dict[str, Any] = {
        "path": path,
        "key_column": key_column,
        "sheet": sheet,
        "total_rows": len(rows),
        "matched_exact": matched_exact,
        "matched_filename": matched_filename,
        "ambiguous": ambiguous[:50],  # Preview, der Rest ist in der Datei
        "ambiguous_total": len(ambiguous),
        "not_found": not_found[:50],
        "not_found_total": len(not_found),
        "warnings": warnings,
    }

    if not commit:
        # Trockenlauf — nur Report
        result["committed"] = False
        result["columns_written"] = []
        result["rows_written"] = 0
        result["hint"] = (
            "Trockenlauf. Mit commit=true werden die Metadaten in "
            "agent_source_metadata geschrieben. Ambiguous/not-found bitte "
            "zuerst klaeren (Pfade in der Begleit-Excel korrigieren oder "
            "ignore-Liste erweitern)."
        )
        return result

    # Commit: schreiben
    metadata_cols = [c for c in header_cols if c not in ignored]
    if not metadata_cols:
        result["committed"] = False
        result["hint"] = "Keine Metadaten-Spalten zum Schreiben (alle ignoriert)."
        return result

    conn = db_connect()
    rows_written = 0
    try:
        for source_id, row in resolved_rows:
            for col in metadata_cols:
                val = row.get(col)
                if val is None or str(val).strip() == "":
                    continue
                # UPSERT per ON CONFLICT (Unique auf source_id+key+source_of_truth)
                conn.execute(
                    """
                    INSERT INTO agent_source_metadata
                        (source_id, key, value, source_of_truth, updated_at)
                    VALUES (?, ?, ?, 'begleit-excel', datetime('now'))
                    ON CONFLICT(source_id, key, source_of_truth) DO UPDATE SET
                        value = excluded.value,
                        updated_at = excluded.updated_at
                    """,
                    (source_id, col, str(val)),
                )
                rows_written += 1
        conn.commit()
    finally:
        conn.close()

    result["committed"] = True
    result["columns_written"] = metadata_cols
    result["rows_written"] = rows_written
    result["hint"] = (
        "Metadaten geschrieben. Abfrage mit: "
        "SELECT s.rel_path, m.key, m.value FROM agent_source_metadata m "
        "JOIN agent_sources s ON s.id = m.source_id WHERE m.source_of_truth='begleit-excel'"
    )
    return result


# ===========================================================================
# sources_detect_duplicates — sha256-Gruppen → duplicate-of Relations
# ===========================================================================


@register(
    name="sources_detect_duplicates",
    description=(
        "Erkennt Duplikate anhand identischer sha256-Hashes und schreibt "
        "'duplicate-of'-Relationen in agent_source_relations. Pro Duplikat-Set "
        "wird der aelteste Eintrag (first_seen_at) zum 'kanonischen' erklaert — "
        "alle anderen erhalten eine 'duplicate-of'-Relation, die auf den "
        "Kanonischen zeigt. Confidence: 1.0 (Hash-Gleichheit ist eindeutig). "
        "Idempotent: Re-Runs ergaenzen nur neue Duplikate, bestehende Relationen "
        "werden nicht dupliziert (unique index)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "min_group_size": {
                "type": "integer",
                "description": (
                    "Mindestgroesse einer Hash-Gruppe, um als Duplikat zu zaehlen. "
                    "Default 2 (jede Mehrfachkopie). 3+ liefert nur echte "
                    "'3-fach-oder-mehr'-Faelle."
                ),
            },
            "include_deleted": {
                "type": "boolean",
                "description": (
                    "Auch Dateien mit status='deleted' einbeziehen (Default false)."
                ),
            },
        },
        "required": [],
    },
    returns=(
        "{scanned, groups_found, new_relations, duplicate_sets: "
        "[{sha256, canonical: {id, rel_path}, copies: [{id, rel_path}]}]}"
    ),
)
def _sources_detect_duplicates(
    *,
    min_group_size: int = 2,
    include_deleted: bool = False,
) -> dict[str, Any]:
    min_group_size = max(2, int(min_group_size))
    status_clause = "" if include_deleted else " AND status = 'active'"

    conn = db_connect()
    try:
        # Existiert die Relations-Tabelle? (Projekt-Template-Migration 002)
        has_rel = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='agent_source_relations'"
        ).fetchone()
        if not has_rel:
            return {
                "error": (
                    "agent_source_relations existiert nicht. Bitte "
                    "'disco project init <slug>' neu laufen lassen, damit die "
                    "Template-Migration 002 angewendet wird."
                ),
            }

        # Hash-Gruppen mit >= min_group_size
        group_rows = conn.execute(
            f"""
            SELECT sha256, GROUP_CONCAT(id) AS ids, COUNT(*) AS n
            FROM agent_sources
            WHERE sha256 IS NOT NULL AND sha256 != ''{status_clause}
            GROUP BY sha256
            HAVING COUNT(*) >= ?
            ORDER BY n DESC, sha256
            """,
            (min_group_size,),
        ).fetchall()

        new_relations = 0
        duplicate_sets: list[dict] = []
        for gr in group_rows:
            ids = [int(x) for x in gr["ids"].split(",")]
            # Kanonisch = der mit frueheste first_seen_at (bei Gleichstand kleinste id)
            rows = conn.execute(
                "SELECT id, rel_path, first_seen_at FROM agent_sources WHERE id IN "
                "(" + ",".join("?" * len(ids)) + ") "
                "ORDER BY first_seen_at ASC, id ASC",
                ids,
            ).fetchall()
            canonical = rows[0]
            copies = rows[1:]

            for cp in copies:
                # INSERT OR IGNORE — idempotent dank UNIQUE-Index
                cur = conn.execute(
                    """
                    INSERT OR IGNORE INTO agent_source_relations
                        (from_source_id, to_source_id, kind, confidence,
                         detected_by, detected_at, note)
                    VALUES (?, ?, 'duplicate-of', 1.0, 'duplicate-hash',
                            datetime('now'), ?)
                    """,
                    (cp["id"], canonical["id"],
                     f"sha256={gr['sha256'][:16]}..."),
                )
                if cur.rowcount:
                    new_relations += 1

            duplicate_sets.append({
                "sha256": gr["sha256"],
                "size_in_set": gr["n"],
                "canonical": {"id": canonical["id"], "rel_path": canonical["rel_path"]},
                "copies": [{"id": c["id"], "rel_path": c["rel_path"]} for c in copies],
            })

        conn.commit()

        # Uebersicht: wie viele Dateien sind nicht-kanonisch (koennten entfernt werden)?
        dup_count = conn.execute(
            """
            SELECT COUNT(DISTINCT from_source_id) FROM agent_source_relations
            WHERE kind = 'duplicate-of'
            """
        ).fetchone()[0]
        scanned = conn.execute(
            f"SELECT COUNT(*) FROM agent_sources WHERE 1=1{status_clause}"
        ).fetchone()[0]
    finally:
        conn.close()

    return {
        "scanned": scanned,
        "groups_found": len(duplicate_sets),
        "new_relations": new_relations,
        "non_canonical_total": dup_count,
        "duplicate_sets": duplicate_sets[:20],
        "duplicate_sets_total": len(duplicate_sets),
        "truncated": len(duplicate_sets) > 20,
        "hint": (
            "Non-canonical Dateien koennen bei Bedarf gesichtet werden: "
            "SELECT s.rel_path AS kopie, c.rel_path AS kanonisch "
            "FROM agent_source_relations r "
            "JOIN agent_sources s ON s.id = r.from_source_id "
            "JOIN agent_sources c ON c.id = r.to_source_id "
            "WHERE r.kind='duplicate-of' ORDER BY c.rel_path, s.rel_path"
        ),
    }
