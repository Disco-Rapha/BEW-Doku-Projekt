"""Server-seitige Import-/Export-Tools fuer Tabellen-Daten (Phase 2b-Erweiterung).

Hintergrund: Excel/CSV-Imports per Code Interpreter + base64-Bridging sind
unzuverlaessig (GPT-5 verstuemmelt grosse base64-Strings beim Transfer in
den CI-Block). Diese Tools loesen das Problem, indem die Imports
**komplett server-seitig** ablaufen — der Agent uebergibt nur Pfad +
Ziel-Tabelle, das Backend liest die Datei und schreibt direkt in die DB.

Drei Funktionen:
  - xlsx_inspect: kurze Inspektion einer .xlsx (Sheets, Header, Zeilen)
  - import_xlsx_to_table: Sheet aus .xlsx -> work_*/agent_*-Tabelle
  - import_csv_to_table:  CSV -> work_*/agent_*-Tabelle

Sicherheit:
  - Pfade werden gegen settings.data_dir resolved (kein Traversal).
  - Ziel-Tabellen MUESSEN mit 'work_' oder 'agent_' beginnen.
  - Default: existierende Tabellen werden NICHT ueberschrieben (drop_existing=False)
    -> der Agent muss bewusst drop_existing=True setzen.
"""

from __future__ import annotations

import csv
import io
import re
import sqlite3
from pathlib import Path
from typing import Any

from ...db import connect
from . import register
from .data import AGENT_NAMESPACE_PREFIXES
from .fs import _resolve_under_data, _data_root


# ---------------------------------------------------------------------------
# Konstanten
# ---------------------------------------------------------------------------

MAX_ROWS_PER_IMPORT = 50_000          # bei mehr Zeilen Hinweis auf Job-System
MAX_COLUMNS_PER_TABLE = 80
DEFAULT_BATCH_SIZE = 500


# ---------------------------------------------------------------------------
# Helfer
# ---------------------------------------------------------------------------


_NON_ID = re.compile(r"[^a-z0-9_]+")
_LEAD_DIGIT = re.compile(r"^\d+")


def _to_snake_case(label: str) -> str:
    """Wandelt einen Excel-Header in einen sauberen SQL-Spaltennamen."""
    s = (label or "").strip().lower()
    # Umlaute und ß als ASCII-Naeherung
    s = (s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
           .replace("ß", "ss"))
    # Nicht-Identifier-Zeichen -> Underscore
    s = _NON_ID.sub("_", s)
    s = s.strip("_")
    if not s:
        s = "col"
    # SQL-Spalten duerfen nicht mit Ziffer beginnen
    if _LEAD_DIGIT.match(s):
        s = "c_" + s
    return s


def _check_target_table(name: str) -> str:
    """Validiert Tabellen-Namen: muss work_ oder agent_ Prefix haben."""
    n = (name or "").strip()
    if not n or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", n):
        raise ValueError(f"Ungueltiger Tabellenname: {name!r}")
    if not n.lower().startswith(AGENT_NAMESPACE_PREFIXES):
        raise ValueError(
            f"Tabellenname muss mit {list(AGENT_NAMESPACE_PREFIXES)} beginnen, "
            f"nicht: {n!r}."
        )
    return n


def _resolve_columns(headers: list[str], rename: dict[str, str] | None) -> list[str]:
    """Wendet rename-Map an, sonst snake_case-Konvertierung. Macht Namen unique."""
    rename = rename or {}
    out: list[str] = []
    seen: dict[str, int] = {}
    for h in headers:
        target = rename.get(h, _to_snake_case(h))
        if target in seen:
            seen[target] += 1
            target = f"{target}_{seen[target]}"
        else:
            seen[target] = 1
        out.append(target)
    return out


def _create_and_insert(
    conn: sqlite3.Connection,
    target_table: str,
    columns: list[str],
    rows: list[list[Any]],
    drop_existing: bool,
    add_id_column: bool,
) -> dict[str, Any]:
    """Erstellt die Tabelle und fuehrt Bulk-INSERT in Batches aus."""
    if drop_existing:
        conn.execute(f"DROP TABLE IF EXISTS {target_table}")

    # Schema: alle Spalten als TEXT — sicher, vom Agent leicht zu queryen,
    # spaeter per CAST konvertierbar.
    col_defs = []
    if add_id_column:
        col_defs.append("id INTEGER PRIMARY KEY AUTOINCREMENT")
    col_defs.extend(f'"{c}" TEXT' for c in columns)
    create_sql = f"CREATE TABLE IF NOT EXISTS {target_table} (\n  {','.join(col_defs)}\n)"
    conn.execute(create_sql)

    placeholders = ", ".join(["?"] * len(columns))
    column_list = ", ".join(f'"{c}"' for c in columns)
    insert_sql = f"INSERT INTO {target_table} ({column_list}) VALUES ({placeholders})"

    inserted = 0
    for i in range(0, len(rows), DEFAULT_BATCH_SIZE):
        batch = rows[i : i + DEFAULT_BATCH_SIZE]
        # Werte in Strings konvertieren (TEXT-Spalten)
        conn.executemany(
            insert_sql,
            [[("" if v is None else str(v)) for v in r] for r in batch],
        )
        inserted += len(batch)

    conn.commit()
    return {
        "target_table": target_table,
        "columns": ([("id",)] if add_id_column else []) + [(c,) for c in columns],
        "columns_flat": (["id"] if add_id_column else []) + columns,
        "rows_inserted": inserted,
        "create_sql": create_sql,
    }


# ===========================================================================
# xlsx_inspect
# ===========================================================================


@register(
    name="xlsx_inspect",
    description=(
        "Inspiziert eine Excel-Datei (.xlsx) unter data/: liefert Sheets-Liste, "
        "je Sheet Anzahl Zeilen/Spalten und die ersten 2-3 Zeilen als Vorschau. "
        "Nutze das, um vor einem Import zu verstehen, welche Sheets es gibt "
        "und wo die Header-Zeile sitzt. Schnell und billig — keine DB-Schreibung."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Pfad zur .xlsx, relativ zu data/."},
            "preview_rows": {
                "type": "integer",
                "description": "Anzahl Zeilen pro Sheet zur Vorschau (Default 3, Max 10).",
            },
        },
        "required": ["path"],
    },
    returns="{path, sheets: [{name, max_row, max_column, preview: [[...], ...]}]}",
)
def _xlsx_inspect(*, path: str, preview_rows: int = 3) -> dict[str, Any]:
    target = _resolve_under_data(path)
    if not target.exists():
        raise ValueError(f"Datei nicht gefunden: {path!r}")
    if target.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Keine .xlsx-Datei: {target.suffix!r}")

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl fehlt. `uv sync` laufen lassen.") from exc

    wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
    n_preview = max(1, min(int(preview_rows or 3), 10))
    out_sheets: list[dict[str, Any]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        preview: list[list[Any]] = []
        # read_only braucht iter_rows
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i > n_preview:
                break
            preview.append(list(row))
        out_sheets.append(
            {
                "name": sn,
                "max_row": max_row,
                "max_column": max_col,
                "preview": preview,
            }
        )
    wb.close()
    return {
        "path": str(target.relative_to(_data_root())),
        "sheets": out_sheets,
    }


# ===========================================================================
# import_xlsx_to_table
# ===========================================================================


@register(
    name="import_xlsx_to_table",
    description=(
        "Importiert ein Sheet einer Excel-Datei direkt in eine work_/agent_-"
        "Tabelle der DB. Server-seitig — der Agent muss nichts im Code "
        "Interpreter jonglieren. "
        "Standardverhalten: Spalten werden zu snake_case (Umlaute zu ae/oe/ue/ss). "
        "Mit columns_rename kann man einzelne Spalten gezielt umbenennen. "
        "drop_existing=true loescht die Tabelle vorher; default ist Append nur "
        "wenn die Tabelle noch nicht existiert (sonst Fehler). "
        "Default add_id=true fuegt eine zusaetzliche id-Spalte (INTEGER PK AUTOINC) ein."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Pfad zur .xlsx, relativ zu data/."},
            "sheet_name": {
                "type": "string",
                "description": "Sheet-Name aus xlsx_inspect.",
            },
            "target_table": {
                "type": "string",
                "description": "Zieltabelle, MUSS mit 'work_' oder 'agent_' beginnen.",
            },
            "header_row": {
                "type": "integer",
                "description": "1-basierte Zeile mit Header (Default 1).",
            },
            "columns_rename": {
                "type": "object",
                "description": (
                    "Optionale Map {original_header: zielname}. Nicht aufgefuehrte "
                    "Spalten werden automatisch in snake_case umgewandelt."
                ),
                "additionalProperties": {"type": "string"},
            },
            "drop_existing": {
                "type": "boolean",
                "description": "Bestehende Tabelle vorher droppen (Default false).",
            },
            "add_id": {
                "type": "boolean",
                "description": "id INTEGER PK AUTOINCREMENT als erste Spalte ergaenzen (Default true).",
            },
            "skip_empty_rows": {
                "type": "boolean",
                "description": "Komplett leere Zeilen ueberspringen (Default true).",
            },
        },
        "required": ["path", "sheet_name", "target_table"],
    },
    returns="{target_table, columns_flat, rows_inserted, sample_row, total_rows}",
)
def _import_xlsx_to_table(
    *,
    path: str,
    sheet_name: str,
    target_table: str,
    header_row: int = 1,
    columns_rename: dict[str, str] | None = None,
    drop_existing: bool = False,
    add_id: bool = True,
    skip_empty_rows: bool = True,
) -> dict[str, Any]:
    table = _check_target_table(target_table)
    target = _resolve_under_data(path)
    if not target.exists():
        raise ValueError(f"Datei nicht gefunden: {path!r}")
    if target.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Keine .xlsx-Datei: {target.suffix!r}")

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl fehlt. `uv sync` laufen lassen.") from exc

    wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' nicht gefunden. Vorhandene Sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    header_row = max(1, int(header_row or 1))

    headers: list[str] = []
    data_rows: list[list[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < header_row:
            continue
        if i == header_row:
            headers = [str(v) if v is not None else f"col_{idx+1}"
                       for idx, v in enumerate(row)]
            continue
        # Datenzeile
        if skip_empty_rows and all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        data_rows.append(list(row))
        if len(data_rows) > MAX_ROWS_PER_IMPORT:
            wb.close()
            raise ValueError(
                f"Mehr als {MAX_ROWS_PER_IMPORT} Zeilen — bitte Sheet aufteilen "
                f"oder Worker-Job nutzen (Phase 2c)."
            )
    wb.close()

    if not headers:
        raise ValueError(f"Header-Zeile {header_row} ist leer in Sheet '{sheet_name}'.")
    if len(headers) > MAX_COLUMNS_PER_TABLE:
        raise ValueError(
            f"Zu viele Spalten ({len(headers)}); Limit {MAX_COLUMNS_PER_TABLE}."
        )

    columns = _resolve_columns(headers, columns_rename)
    # Auf Original-Spaltenanzahl zuschneiden, falls kuerzere Datenzeilen
    n_cols = len(columns)
    data_rows = [
        (r + [None] * (n_cols - len(r))) if len(r) < n_cols else r[:n_cols]
        for r in data_rows
    ]

    # Wenn Tabelle existiert + drop_existing=False → klare Fehlermeldung
    conn = connect()
    try:
        cur = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        )
        if cur.fetchone() and not drop_existing:
            raise ValueError(
                f"Tabelle '{table}' existiert bereits. Setze drop_existing=true "
                f"um sie zu ersetzen."
            )

        result = _create_and_insert(
            conn, table, columns, data_rows,
            drop_existing=drop_existing, add_id_column=add_id,
        )
        # Sample-Zeile fuer den Agent
        sample_cur = conn.execute(f"SELECT * FROM {table} LIMIT 1")
        sample_cols = [d[0] for d in sample_cur.description]
        sample_row = sample_cur.fetchone()
        sample = dict(zip(sample_cols, sample_row)) if sample_row else None

        total_cur = conn.execute(f"SELECT COUNT(*) FROM {table}")
        total = total_cur.fetchone()[0]
    finally:
        conn.close()

    return {
        "target_table": table,
        "source_path": str(target.relative_to(_data_root())),
        "source_sheet": sheet_name,
        "header_row": header_row,
        "columns_flat": result["columns_flat"],
        "rows_inserted": result["rows_inserted"],
        "total_rows_in_table": total,
        "sample_row": sample,
    }


# ===========================================================================
# import_csv_to_table
# ===========================================================================


@register(
    name="import_csv_to_table",
    description=(
        "Importiert eine CSV-Datei direkt in eine work_/agent_-Tabelle. "
        "Server-seitig, schnell, kein Code Interpreter noetig. "
        "Default-Delimiter ist ',' — fuer deutsche Excel-Exporte oft ';'. "
        "encoding default 'utf-8' (BOM wird automatisch entfernt)."
    ),
    parameters={
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "Pfad zur .csv unter data/."},
            "target_table": {
                "type": "string",
                "description": "Zieltabelle, MUSS mit 'work_' oder 'agent_' beginnen.",
            },
            "delimiter": {
                "type": "string",
                "description": "Trennzeichen (Default ','). Bei deutschen Excel-Exporten ';'.",
            },
            "encoding": {
                "type": "string",
                "description": "Encoding (Default 'utf-8'; BOM wird automatisch erkannt).",
            },
            "header_row": {
                "type": "integer",
                "description": "1-basierte Zeile mit Header (Default 1).",
            },
            "columns_rename": {
                "type": "object",
                "description": "Optionale Map {original_header: zielname}.",
                "additionalProperties": {"type": "string"},
            },
            "drop_existing": {
                "type": "boolean",
                "description": "Bestehende Tabelle vorher droppen (Default false).",
            },
            "add_id": {
                "type": "boolean",
                "description": "id INTEGER PK AUTOINCREMENT als erste Spalte (Default true).",
            },
            "skip_empty_rows": {
                "type": "boolean",
                "description": "Leere Zeilen ueberspringen (Default true).",
            },
        },
        "required": ["path", "target_table"],
    },
    returns="{target_table, columns_flat, rows_inserted, sample_row, total_rows_in_table}",
)
def _import_csv_to_table(
    *,
    path: str,
    target_table: str,
    delimiter: str = ",",
    encoding: str = "utf-8",
    header_row: int = 1,
    columns_rename: dict[str, str] | None = None,
    drop_existing: bool = False,
    add_id: bool = True,
    skip_empty_rows: bool = True,
) -> dict[str, Any]:
    table = _check_target_table(target_table)
    target = _resolve_under_data(path)
    if not target.exists():
        raise ValueError(f"Datei nicht gefunden: {path!r}")

    # Datei lesen, BOM ggf. entfernen
    raw = target.read_bytes()
    try:
        text = raw.decode(encoding)
    except UnicodeDecodeError as exc:
        raise ValueError(f"Encoding-Fehler ({encoding}): {exc}") from exc
    if text.startswith("\ufeff"):
        text = text[1:]

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise ValueError("CSV ist leer.")

    header_row = max(1, int(header_row or 1))
    if len(rows) < header_row:
        raise ValueError(f"CSV hat nur {len(rows)} Zeilen, header_row={header_row}.")

    headers = [h if h else f"col_{i+1}" for i, h in enumerate(rows[header_row - 1])]
    data_raw = rows[header_row:]
    data: list[list[Any]] = []
    for r in data_raw:
        if skip_empty_rows and all((not c or not c.strip()) for c in r):
            continue
        data.append(list(r))
        if len(data) > MAX_ROWS_PER_IMPORT:
            raise ValueError(
                f"Mehr als {MAX_ROWS_PER_IMPORT} Zeilen — Worker-Job (Phase 2c) nutzen."
            )

    if len(headers) > MAX_COLUMNS_PER_TABLE:
        raise ValueError(f"Zu viele Spalten ({len(headers)})")

    columns = _resolve_columns(headers, columns_rename)
    n_cols = len(columns)
    data = [
        (r + [None] * (n_cols - len(r))) if len(r) < n_cols else r[:n_cols]
        for r in data
    ]

    conn = connect()
    try:
        cur = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table,),
        )
        if cur.fetchone() and not drop_existing:
            raise ValueError(
                f"Tabelle '{table}' existiert bereits. Setze drop_existing=true."
            )
        result = _create_and_insert(
            conn, table, columns, data,
            drop_existing=drop_existing, add_id_column=add_id,
        )
        sample_cur = conn.execute(f"SELECT * FROM {table} LIMIT 1")
        sample_cols = [d[0] for d in sample_cur.description]
        sample_row = sample_cur.fetchone()
        sample = dict(zip(sample_cols, sample_row)) if sample_row else None
        total_cur = conn.execute(f"SELECT COUNT(*) FROM {table}")
        total = total_cur.fetchone()[0]
    finally:
        conn.close()

    return {
        "target_table": table,
        "source_path": str(target.relative_to(_data_root())),
        "delimiter": delimiter,
        "header_row": header_row,
        "columns_flat": result["columns_flat"],
        "rows_inserted": result["rows_inserted"],
        "total_rows_in_table": total,
        "sample_row": sample,
    }


# ===========================================================================
# build_xlsx_from_tables  (Cowork-Style Excel-Export, server-seitig)
# ===========================================================================


# Style-Definitionen, einmalig fuer den Server
def _excel_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "HEADER_FILL":  PatternFill("solid", fgColor="305496"),
        "HEADER_FONT":  Font(bold=True, color="FFFFFF", size=11),
        "HEADER_ALIGN": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "ZEBRA_FILL":   PatternFill("solid", fgColor="F2F2F2"),
        "BORDER":       Border(*[Side(style="thin", color="BFBFBF")] * 4),
        "STATUS": {
            "Erfuellt":  PatternFill("solid", fgColor="C6EFCE"),
            "Erfüllt":   PatternFill("solid", fgColor="C6EFCE"),
            "Teilweise": PatternFill("solid", fgColor="FFEB9C"),
            "Fehlend":   PatternFill("solid", fgColor="FFC7CE"),
            "Pruefen":   PatternFill("solid", fgColor="DDEBF7"),
            "Prüfen":    PatternFill("solid", fgColor="DDEBF7"),
        },
        "LINK_FONT":    Font(color="0563C1", underline="single"),
    }


def _autosize_columns(ws, headers: list[str], max_widths: dict[int, int]) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx, h in enumerate(headers, start=1):
        cur = max_widths.get(col_idx, len(str(h)))
        width = max(8, min(80, cur + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_one_sheet(wb, spec: dict, fetched_rows: list[dict], styles: dict) -> dict:
    """Baut ein einzelnes Sheet. Liefert Statistik zurueck."""
    from openpyxl.utils import get_column_letter

    name = spec["name"]
    sheet = wb.create_sheet(name[:31])  # Excel-Limit: 31 Zeichen

    # Spalten/Header bestimmen
    if not fetched_rows:
        # Wenn leer: nur Header aus column_renames oder von select_columns
        select_cols = spec.get("select_columns") or []
        renames = spec.get("column_renames") or {}
        headers = [renames.get(c, c) for c in select_cols] or ["(leer)"]
        col_keys = select_cols
    else:
        select_cols = spec.get("select_columns") or list(fetched_rows[0].keys())
        renames = spec.get("column_renames") or {}
        headers = [renames.get(c, c) for c in select_cols]
        col_keys = select_cols

    # Header schreiben
    for col, h in enumerate(headers, 1):
        c = sheet.cell(row=1, column=col, value=h)
        c.fill = styles["HEADER_FILL"]
        c.font = styles["HEADER_FONT"]
        c.alignment = styles["HEADER_ALIGN"]
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"

    # Spalten-Index der Status-/Hyperlink-Spalten (1-basiert)
    status_col = spec.get("status_column")
    hyperlink_col = spec.get("hyperlink_column")
    status_idx = (col_keys.index(status_col) + 1) if status_col in col_keys else None
    hyperlink_idx = (col_keys.index(hyperlink_col) + 1) if hyperlink_col in col_keys else None

    # Datenzeilen + Spaltenbreiten-Tracking
    max_widths: dict[int, int] = {i: len(str(h)) for i, h in enumerate(headers, 1)}
    for r_idx, row_dict in enumerate(fetched_rows, start=2):
        for c_idx, key in enumerate(col_keys, start=1):
            val = row_dict.get(key)
            cell = sheet.cell(row=r_idx, column=c_idx, value=val)
            cell.border = styles["BORDER"]
            if r_idx % 2 == 0:
                cell.fill = styles["ZEBRA_FILL"]
            # Status-Faerbung
            if status_idx and c_idx == status_idx and val in styles["STATUS"]:
                cell.fill = styles["STATUS"][val]
            # Hyperlinks: Format "Anzeige|#Sheet!A1"
            if hyperlink_idx and c_idx == hyperlink_idx and val and "|" in str(val):
                text, target = str(val).split("|", 1)
                cell.value = text
                cell.hyperlink = target
                cell.font = styles["LINK_FONT"]
                val = text
            # Width-Tracking
            s = "" if val is None else str(val)
            if len(s) > max_widths.get(c_idx, 0):
                max_widths[c_idx] = len(s)

    # AutoFilter auf gesamten Datenbereich
    if fetched_rows:
        last_col = get_column_letter(len(headers))
        sheet.auto_filter.ref = f"A1:{last_col}{len(fetched_rows) + 1}"

    _autosize_columns(sheet, headers, max_widths)

    return {
        "sheet_name": sheet.title,
        "row_count": len(fetched_rows),
        "column_count": len(headers),
        "headers": headers,
    }


@register(
    name="build_xlsx_from_tables",
    description=(
        "Erzeugt eine professionell formatierte Excel direkt server-seitig — "
        "Multi-Sheet, Header-Style, Spaltenbreite, AutoFilter, Freeze Panes, "
        "optional Status-Zellfarben (gruen/gelb/rot) und Hyperlinks zwischen Sheets. "
        "Du gibst nur eine Spec: pro Sheet ein SQL-SELECT (oder fertige rows), "
        "optional Spalten-Umbenennungen, optional Status-Spalte. "
        "Vorteile: deterministisch, schnell (Sekunden), funktioniert fuer beliebig "
        "grosse Excels (10 MB+), kein base64-Bridging notwendig."
    ),
    parameters={
        "type": "object",
        "properties": {
            "target_path": {
                "type": "string",
                "description": (
                    "Zielpfad relativ zu data/ (z.B. 'exports/ibl_2026-04-16_v1.xlsx'). "
                    "Muss mit .xlsx enden. Kein Ueberschreiben — wenn die Datei "
                    "schon existiert, Fehler (nutze Versions-Suffix)."
                ),
            },
            "title": {
                "type": "string",
                "description": "Titel oben im Uebersichts-Sheet (z.B. 'IBL Lagerhalle Reuter').",
            },
            "overview_rows": {
                "type": "array",
                "description": (
                    "Optional: Liste von [Kennzahl, Wert]-Paaren fuer das Uebersichts-Sheet "
                    "(z.B. [['Komponenten', 322], ['IBL', 72]]). Wenn leer, wird kein Uebersichts-Sheet erzeugt."
                ),
                "items": {
                    "type": "array",
                    "items": {"type": ["string", "number", "null"]},
                },
            },
            "sheets": {
                "type": "array",
                "description": (
                    "Liste der Daten-Sheets. Jedes Sheet ist ein Objekt mit:\n"
                    "  name (str, max 31 Zeichen)\n"
                    "  sql (str, READ-ONLY SELECT) ODER rows (list[dict] mit gleichen Keys)\n"
                    "  select_columns (list[str], optional — Reihenfolge/Auswahl der Spalten; "
                    "Default: alle aus dem ersten Result-Dict)\n"
                    "  column_renames (object, optional — {original_key: angezeigter_header})\n"
                    "  status_column (str, optional — welcher key 'Erfuellt'/'Teilweise'/'Fehlend' enthaelt)\n"
                    "  hyperlink_column (str, optional — Werte im Format 'Anzeige|#OtherSheet!A1')"
                ),
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string"},
                        "sql": {"type": "string"},
                        "rows": {
                            "type": "array",
                            "items": {"type": "object"},
                        },
                        "select_columns": {"type": "array", "items": {"type": "string"}},
                        "column_renames": {
                            "type": "object",
                            "additionalProperties": {"type": "string"},
                        },
                        "status_column": {"type": "string"},
                        "hyperlink_column": {"type": "string"},
                    },
                    "required": ["name"],
                },
            },
        },
        "required": ["target_path", "sheets"],
    },
    returns="{path, total_size, sheets: [{sheet_name, row_count, column_count, headers}]}",
)
def _build_xlsx_from_tables(
    *,
    target_path: str,
    sheets: list[dict],
    title: str = "Report",
    overview_rows: list[list] | None = None,
) -> dict[str, Any]:
    if not target_path.lower().endswith(".xlsx"):
        raise ValueError("target_path muss auf .xlsx enden.")
    target = _resolve_under_data(target_path)
    if target.exists():
        raise ValueError(
            f"Datei existiert bereits: {target_path}. Versions-Suffix nutzen "
            f"(z.B. _v2.xlsx)."
        )
    target.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl fehlt. `uv sync` laufen lassen.") from exc

    if not sheets:
        raise ValueError("Mindestens ein Sheet ist erforderlich.")

    styles = _excel_styles()
    wb = Workbook()
    wb.remove(wb.active)

    # Optional: Uebersichts-Sheet
    if overview_rows:
        ov = wb.create_sheet("0-Übersicht")
        ov["A1"] = title
        ov["A1"].font = Font(bold=True, size=14)
        from datetime import date
        ov["A2"] = f"Stand: {date.today().isoformat()}"
        ov["A2"].font = Font(italic=True, color="555555")
        # Header-Zeile (4) fuer Kennzahlen
        for c_idx, h in enumerate(["Kennzahl", "Wert"], 1):
            cell = ov.cell(row=4, column=c_idx, value=h)
            cell.fill = styles["HEADER_FILL"]
            cell.font = styles["HEADER_FONT"]
        for r_idx, kv in enumerate(overview_rows, start=5):
            if len(kv) < 2:
                continue
            ov.cell(row=r_idx, column=1, value=kv[0]).border = styles["BORDER"]
            ov.cell(row=r_idx, column=2, value=kv[1]).border = styles["BORDER"]
        ov.column_dimensions["A"].width = 32
        ov.column_dimensions["B"].width = 18

    # Datensheets bauen — pro Sheet ggf. SQL ausfuehren
    sheet_stats = []
    conn = connect()
    try:
        for spec in sheets:
            name = spec.get("name")
            if not name:
                raise ValueError("Sheet-Spec ohne 'name'.")
            # Daten holen
            if "rows" in spec and spec["rows"] is not None:
                fetched = list(spec["rows"])
            elif "sql" in spec and spec["sql"]:
                # Nur SELECT erlaubt
                from .data import _check_read_only
                _check_read_only(spec["sql"])
                cur = conn.execute(spec["sql"])
                cols = [d[0] for d in cur.description]
                fetched = [dict(zip(cols, r)) for r in cur.fetchall()]
            else:
                raise ValueError(f"Sheet '{name}' braucht 'sql' oder 'rows'.")
            stat = _build_one_sheet(wb, spec, fetched, styles)
            sheet_stats.append(stat)
    finally:
        conn.close()

    wb.save(target)
    size = target.stat().st_size

    return {
        "path": str(target.relative_to(_data_root())),
        "total_size": size,
        "sheets": sheet_stats,
    }
